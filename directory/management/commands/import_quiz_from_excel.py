# -*- coding: utf-8 -*-
# directory/management/commands/import_quiz_from_excel.py
"""
Django management команда для импорта вопросов из Excel (.xls, .xlsx)

Формат таблицы:
┌────┬──────────┬───────────┬───────────┬───────────┬───────────┐
│ №  │ Вопрос   │ Вариант 1 │ Вариант 2 │ Вариант 3 │ Вариант 4 │
├────┼──────────┼───────────┼───────────┼───────────┼───────────┤
│ 1  │ Текст?   │ Ответ А   │ Ответ Б   │ Ответ В   │ Ответ Г   │
│ 2  │ Текст?   │ Ответ 1   │ Ответ 2   │ Ответ 3   │ Ответ 4   │
└────┴──────────┴───────────┴───────────┴───────────┴───────────┘

Правильный ответ выделен ЖИРНЫМ шрифтом.

Использование:
    # Импорт только вопросов
    python manage.py import_quiz_from_excel <файл.xlsx> --category="Название раздела"

    # Импорт вопросов с изображениями
    python manage.py import_quiz_from_excel <файл.xlsx> --category="Раздел" --images-dir="путь/к/папке"

Опции:
    --category: Название раздела для импорта (обязательно)
    --images-dir: Папка с изображениями (01.jpg, 02.png и т.д.)
    --dry-run: Тестовый запуск без сохранения в БД
    --clear: Очистить существующие вопросы в разделе
    --skip-header: Пропустить первую строку (заголовок таблицы)

Примеры:
    # Импорт с очисткой и изображениями
    py manage.py import_quiz_from_excel questions.xlsx --category="ОТ" --images-dir="картинки" --clear
"""

from django.core.management.base import BaseCommand, CommandError
from directory.models import QuizCategory, Question, Answer
import openpyxl
from openpyxl.styles import Font
import xlrd
from pathlib import Path
from django.core.files.base import ContentFile


class Command(BaseCommand):
    help = 'Импорт вопросов экзамена из Excel (.xls, .xlsx)'

    def add_arguments(self, parser):
        parser.add_argument(
            'excel_file',
            type=str,
            help='Путь к Excel файлу (.xls или .xlsx)'
        )
        parser.add_argument(
            '--category',
            type=str,
            required=True,
            help='Название раздела для импорта вопросов'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Тестовый запуск без сохранения в БД'
        )
        parser.add_argument(
            '--clear',
            action='store_true',
            help='Очистить существующие вопросы в разделе'
        )
        parser.add_argument(
            '--skip-header',
            action='store_true',
            default=True,
            help='Пропустить первую строку (заголовок)'
        )
        parser.add_argument(
            '--images-dir',
            type=str,
            help='Путь к папке с изображениями (1.jpg, 2.png и т.д.)'
        )

    def handle(self, *args, **options):
        excel_file = options['excel_file']
        category_name = options['category']
        dry_run = options['dry_run']
        clear = options['clear']
        skip_header = options['skip_header']
        images_dir = options.get('images_dir')

        # Настройка кодировки для вывода
        import sys
        if sys.platform == 'win32':
            import io
            sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

        self.stdout.write(self.style.SUCCESS(f'\n{"="*70}'))
        self.stdout.write(self.style.SUCCESS('Импорт вопросов из Excel'))
        self.stdout.write(self.style.SUCCESS(f'{"="*70}\n'))

        # Получаем или создаем раздел
        if not dry_run:
            category, created = QuizCategory.objects.get_or_create(
                name=category_name,
                defaults={'description': ''}
            )
            if created:
                self.stdout.write(self.style.SUCCESS(f'[+] Создан новый раздел: {category_name}'))
            else:
                self.stdout.write(self.style.WARNING(f'[*] Используется существующий раздел: {category_name}'))

            # Очистка существующих вопросов
            if clear:
                deleted_count = Question.objects.filter(category=category).delete()[0]
                self.stdout.write(self.style.WARNING(f'[-] Удалено {deleted_count} существующих вопросов'))
        else:
            self.stdout.write(self.style.WARNING('ТЕСТОВЫЙ РЕЖИМ (dry-run) - данные не будут сохранены'))
            category = None

        # Определяем тип файла и парсим
        if excel_file.endswith('.xlsx'):
            questions = self.parse_xlsx(excel_file, skip_header)
        elif excel_file.endswith('.xls'):
            questions = self.parse_xls(excel_file, skip_header)
        else:
            raise CommandError('Неподдерживаемый формат файла. Используйте .xls или .xlsx')

        self.stdout.write(self.style.SUCCESS(f'\n[OK] Распознано вопросов: {len(questions)}\n'))

        # Статистика
        questions_with_correct = sum(1 for q in questions if any(a['is_correct'] for a in q['answers']))
        self.stdout.write(f'  • С отмеченными правильными ответами: {questions_with_correct}\n')

        # Импортируем вопросы
        imported_count = 0
        warnings = []
        errors = []

        for idx, question_data in enumerate(questions, 1):
            try:
                # Проверка: есть ли правильный ответ?
                has_correct = any(a['is_correct'] for a in question_data['answers'])
                if not has_correct:
                    warnings.append(f'Вопрос #{idx}: нет правильного ответа (жирного текста)')

                if dry_run:
                    # Просто выводим информацию
                    self.stdout.write(f'\n[{idx}] {question_data["text"][:80]}...')
                    self.stdout.write(f'    Вариантов ответов: {len(question_data["answers"])}')

                    for i, answer in enumerate(question_data["answers"], 1):
                        marker = ' [ПРАВИЛЬНЫЙ]' if answer['is_correct'] else ''
                        self.stdout.write(f'      {i}. {answer["text"][:60]}{marker}')
                else:
                    # Создаем вопрос
                    question = Question.objects.create(
                        category=category,
                        question_text=question_data['text'],
                        explanation=question_data.get('explanation', ''),
                        order=idx
                    )

                    # Добавляем изображение если есть папка с изображениями
                    if images_dir:
                        image_path = self.find_image_for_question(images_dir, idx)
                        if image_path:
                            with open(image_path, 'rb') as img_file:
                                import os
                                question.image.save(
                                    os.path.basename(image_path),
                                    ContentFile(img_file.read()),
                                    save=False
                                )

                    # Создаем варианты ответов
                    for answer_data in question_data['answers']:
                        Answer.objects.create(
                            question=question,
                            answer_text=answer_data['text'],
                            is_correct=answer_data['is_correct'],
                            order=answer_data['order']
                        )

                    # Сохраняем вопрос с изображением
                    question.save()
                    imported_count += 1

            except Exception as e:
                error_msg = f'Ошибка при импорте вопроса #{idx}: {e}'
                errors.append(error_msg)
                self.stdout.write(self.style.ERROR(f'[ERROR] {error_msg}'))

        # Итоги
        self.stdout.write(self.style.SUCCESS(f'\n{"="*70}'))
        if dry_run:
            self.stdout.write(self.style.WARNING('ТЕСТОВЫЙ РЕЖИМ - данные не сохранены'))
            self.stdout.write(f'Всего вопросов для импорта: {len(questions)}')
        else:
            self.stdout.write(self.style.SUCCESS(f'[OK] Успешно импортировано вопросов: {imported_count}'))

        if warnings:
            self.stdout.write(self.style.WARNING(f'\n[!] Предупреждений: {len(warnings)}'))
            for warning in warnings[:5]:  # Показываем первые 5
                self.stdout.write(self.style.WARNING(f'  - {warning}'))

        if errors:
            self.stdout.write(self.style.ERROR(f'\n[ERROR] Ошибок: {len(errors)}'))
            for error in errors[:5]:
                self.stdout.write(self.style.ERROR(f'  - {error}'))

        self.stdout.write(self.style.SUCCESS(f'{"="*70}\n'))

    def parse_xlsx(self, filepath, skip_header=True):
        """Парсинг .xlsx файла"""
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active

        questions = []
        start_row = 2 if skip_header else 1

        for row_idx, row in enumerate(sheet.iter_rows(min_row=start_row), start=1):
            # Пропускаем пустые строки
            if not row[1].value:  # Если нет текста вопроса
                continue

            # Колонка B (индекс 1) - вопрос
            question_text = str(row[1].value).strip()

            # Колонки C, D, E, F (индексы 2-5) - варианты ответов
            answers = []
            for i in range(2, 6):  # Колонки C-F
                if i < len(row) and row[i].value:
                    answer_text = str(row[i].value).strip()

                    # Проверяем, жирный ли текст
                    is_bold = False
                    if row[i].font and row[i].font.bold:
                        is_bold = True

                    answers.append({
                        'text': answer_text,
                        'is_correct': is_bold,
                        'order': i - 1  # 1, 2, 3, 4
                    })

            if answers:
                questions.append({
                    'text': question_text,
                    'answers': answers,
                    'explanation': ''
                })

        return questions

    def parse_xls(self, filepath, skip_header=True):
        """Парсинг .xls файла (старый формат Excel)"""
        workbook = xlrd.open_workbook(filepath, formatting_info=True)
        sheet = workbook.sheet_by_index(0)

        questions = []
        start_row = 1 if skip_header else 0

        for row_idx in range(start_row, sheet.nrows):
            # Колонка B (индекс 1) - вопрос
            question_cell = sheet.cell(row_idx, 1)
            if not question_cell.value:
                continue

            question_text = str(question_cell.value).strip()

            # Колонки C, D, E, F (индексы 2-5) - варианты ответов
            answers = []
            for col_idx in range(2, 6):
                if col_idx < sheet.ncols:
                    answer_cell = sheet.cell(row_idx, col_idx)
                    if answer_cell.value:
                        answer_text = str(answer_cell.value).strip()

                        # Проверяем, жирный ли текст в .xls
                        is_bold = False
                        try:
                            xf_index = answer_cell.xf_index
                            xf = workbook.format_map[xf_index]
                            font = workbook.font_list[xf.font_index]
                            is_bold = font.weight >= 700  # 700 = bold
                        except:
                            is_bold = False

                        answers.append({
                            'text': answer_text,
                            'is_correct': is_bold,
                            'order': col_idx - 1
                        })

            if answers:
                questions.append({
                    'text': question_text,
                    'answers': answers,
                    'explanation': ''
                })

        return questions
    def find_image_for_question(self, images_dir, question_number):
        """
        Поиск изображения для вопроса по номеру.
        Формат: 1.jpg, 01.jpg, 001.jpg, 2.png, 02.png и т.д.
        """
        images_path = Path(images_dir)

        if not images_path.exists():
            return None

        # Варианты форматирования номера: 1, 01, 001
        number_formats = [
            str(question_number),              # 1
            f"{question_number:02d}",          # 01
            f"{question_number:03d}",          # 001
        ]

        # Расширения файлов
        extensions = ['.jpg', '.jpeg', '.png', '.gif']

        # Проверяем все комбинации
        for num_format in number_formats:
            for ext in extensions:
                image_file = images_path / f"{num_format}{ext}"
                if image_file.exists():
                    return str(image_file)

        return None
