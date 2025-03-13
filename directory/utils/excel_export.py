# directory/utils/excel_export.py

import os
import tempfile
import logging
from copy import copy

import openpyxl  # Добавляем прямой импорт openpyxl для проверки типов
from django.conf import settings
from django.http import FileResponse, HttpResponseBadRequest
from django.shortcuts import redirect
from django.contrib import messages
from openpyxl import load_workbook

from directory.models import Employee, SIZNorm

# Настройка логгера
logger = logging.getLogger(__name__)


def copy_row_styles(ws, src_row, dst_row):
    """
    Копирует стили из строки src_row в строку dst_row.

    Args:
        ws: Объект Worksheet.
        src_row: Номер строки-источника стилей.
        dst_row: Номер строки-получателя стилей.

    Note:
        Теперь эта функция корректно обрабатывает объединенные ячейки.
    """
    try:
        max_col = ws.max_column
        for col in range(1, max_col + 1):
            src_cell = ws.cell(row=src_row, column=col)
            dst_cell = ws.cell(row=dst_row, column=col)

            # Проверяем, является ли целевая ячейка объединенной
            if isinstance(dst_cell, openpyxl.cell.cell.MergedCell):
                # Пропускаем объединенные ячейки, так как мы не можем напрямую задать им стиль
                continue

            # Проверяем, имеет ли исходная ячейка стиль
            if src_cell.has_style:
                dst_cell._style = copy(src_cell._style)
    except Exception as e:
        logger.error(f"Ошибка при копировании стилей: {str(e)}")
        raise


def safe_set_cell_value(ws, row, col, value):
    """
    Безопасно устанавливает значение в ячейку, обрабатывая объединенные ячейки.

    Args:
        ws: Объект Worksheet.
        row: Номер строки.
        col: Номер столбца.
        value: Значение для установки.
    """
    cell = ws.cell(row=row, column=col)
    # Проверка, является ли ячейка объединенной
    if isinstance(cell, openpyxl.cell.cell.MergedCell):
        # Найдем первую (главную) ячейку объединенной области
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                # Получаем координаты первой ячейки в объединенной области
                top_left = merged_range.min_row, merged_range.min_col
                # Устанавливаем значение в первую ячейку
                ws.cell(row=top_left[0], column=top_left[1]).value = value
                return
        # Если не нашли объединенную область (на всякий случай)
        logger.warning(f"Не удалось найти главную ячейку для объединенной ячейки {cell.coordinate}")
    else:
        # Обычная ячейка - просто устанавливаем значение
        cell.value = value


def insert_data_with_template(ws, data_list, template_row, sign_row, fill_func):
    """
    Вставляет данные из data_list в лист ws, используя template_row как образец.
    Перед блоком подписей (начинающегося с sign_row) вставляются дополнительные строки,
    чтобы не повредить подписи.

    Args:
        ws: Объект Worksheet из openpyxl.
        data_list: Список объектов (например, эталонных норм СИЗ).
        template_row: Номер строки-шаблона (в нашем случае 16-я).
        sign_row: Номер строки, с которой начинается блок подписей (например, 30).
        fill_func: Функция, которая заполняет одну строку данными (ws, row_idx, item).
    """
    count = len(data_list)
    if count == 0:
        logger.info("Нет данных для вставки")
        return

    # Если записей больше одной, вставляем (count-1) строк перед блоком подписей
    if count > 1:
        logger.info(f"Вставка {count - 1} дополнительных строк перед {sign_row}")
        try:
            # Сохраняем информацию о объединенных ячейках до вставки строк
            merged_cells_info = []
            for merged_range in ws.merged_cells.ranges:
                if merged_range.min_row >= sign_row:
                    # Сохраняем настройки объединенных ячеек, которые будут затронуты вставкой
                    merged_cells_info.append((
                        merged_range.min_row,
                        merged_range.min_col,
                        merged_range.max_row,
                        merged_range.max_col
                    ))

            # Вставляем строки
            ws.insert_rows(sign_row, amount=count - 1)

            # Восстанавливаем объединенные ячейки после вставки
            for min_row, min_col, max_row, max_col in merged_cells_info:
                # Сдвигаем на количество вставленных строк
                new_min_row = min_row + (count - 1)
                new_max_row = max_row + (count - 1)
                ws.merge_cells(
                    start_row=new_min_row,
                    start_column=min_col,
                    end_row=new_max_row,
                    end_column=max_col
                )
        except Exception as e:
            logger.error(f"Ошибка при вставке строк: {str(e)}")
            raise

    # Заполняем строки, начиная с шаблонной (template_row)
    for i, item in enumerate(data_list):
        row_idx = template_row + i
        if i > 0:
            try:
                copy_row_styles(ws, template_row, row_idx)
            except Exception as e:
                logger.error(f"Ошибка при копировании стилей для строки {row_idx}: {str(e)}")
                # Продолжаем выполнение, даже если не удалось скопировать стили

        try:
            fill_func(ws, row_idx, item)
        except Exception as e:
            logger.error(f"Ошибка при заполнении строки {row_idx}: {str(e)}")
            raise


def fill_base_norm_row(ws, row_idx, norm):
    """
    Заполняет одну строку эталонной нормы СИЗ.

    Колонки в таблице:
      - Колонка 1: Наименование СИЗ
      - Колонка 2: Классификация
      - Колонка 3: Единица измерения
      - Колонка 4: Количество
      - Колонка 5: Срок носки (если wear_period == 0, выводим "До износа")

    Args:
        ws: Объект Worksheet.
        row_idx: Номер строки для заполнения.
        norm: Объект SIZNorm с данными о норме СИЗ.
    """
    siz = norm.siz

    # Проверяем наличие всех необходимых атрибутов
    name = getattr(siz, 'name', "")
    classification = getattr(siz, 'classification', "")
    unit = getattr(siz, 'unit', "")
    quantity = getattr(norm, 'quantity', "")
    wear_period = getattr(siz, 'wear_period', None)

    # Заполняем ячейки с использованием безопасной функции
    safe_set_cell_value(ws, row_idx, 1, name)
    safe_set_cell_value(ws, row_idx, 2, classification)
    safe_set_cell_value(ws, row_idx, 3, unit)
    safe_set_cell_value(ws, row_idx, 4, quantity)

    # Проверяем wear_period и форматируем значение
    if wear_period == 0:
        safe_set_cell_value(ws, row_idx, 5, "До износа")
    elif wear_period is not None:
        safe_set_cell_value(ws, row_idx, 5, wear_period)
    else:
        safe_set_cell_value(ws, row_idx, 5, "")


def generate_card_excel(request, employee_id):
    """
    Генерирует Excel-файл на основе шаблона Card.xlsx.
    На листе "Лицевая сторона" заполняется шапка карточки и динамически вставляются
    строки с данными эталонных норм СИЗ, начиная с 16-й строки.

    Координаты шапки:
      - ФИО: B5
      - Пол: E5
      - Структурное подразделение/отдел: B7
      - Профессия (должность): B9
      - Рост: E6
      - Размер одежды: E8
      - Размер обуви: E9

    Args:
        request: Объект HttpRequest для возможности добавления сообщений.
        employee_id: Идентификатор сотрудника.

    Returns:
        FileResponse с заполненным Excel-файлом или HttpResponseBadRequest в случае ошибки.
    """
    logger.info(f"Генерация Excel-карточки для сотрудника {employee_id}")
    logger.info(f"BASE_DIR: {settings.BASE_DIR}")
    tmp_file = None

    try:
        # Получаем сотрудника
        try:
            employee = Employee.objects.get(pk=employee_id)
            logger.info(f"Сотрудник найден: {employee.full_name_nominative}")
        except Employee.DoesNotExist:
            logger.error(f"Сотрудник с ID {employee_id} не найден")
            if request:
                messages.error(request, f"Сотрудник с ID {employee_id} не найден")
            return HttpResponseBadRequest("Сотрудник не найден")

        # Проверяем наличие должности
        if not hasattr(employee, 'position') or not employee.position:
            logger.error(f"У сотрудника {employee_id} не указана должность")
            if request:
                messages.error(request, "У сотрудника не указана должность")
            return HttpResponseBadRequest("Отсутствует должность")

        logger.info(f"Должность сотрудника: {employee.position.position_name}")

        # Получаем эталонные нормы для должности
        base_norms = SIZNorm.objects.filter(position=employee.position, condition='').select_related('siz')

        norms_count = base_norms.count()
        logger.info(f"Найдено {norms_count} эталонных норм СИЗ")

        if not base_norms.exists():
            logger.warning(f"Для должности {employee.position} не найдены эталонные нормы СИЗ")
            if request:
                messages.warning(request, "Для данной должности не найдены эталонные нормы СИЗ")

        # 🔄 Расширенная проверка путей к шаблону
        # Прямой путь из комментария пользователя
        direct_path = r"D:\YandexDisk\OT_online\templates\excel\Card.xlsx"

        # Преобразовываем BASE_DIR в абсолютный путь, чтобы исключить любые проблемы с относительными путями
        base_dir_abs = os.path.abspath(settings.BASE_DIR)
        logger.info(f"Абсолютный BASE_DIR: {base_dir_abs}")

        # Пробуем различные варианты путей
        possible_paths = [
            direct_path,  # Прямой путь
            os.path.join(base_dir_abs, 'templates', 'excel', 'Card.xlsx'),
            os.path.join(os.path.dirname(base_dir_abs), 'templates', 'excel', 'Card.xlsx'),
            os.path.join(base_dir_abs, '..', 'templates', 'excel', 'Card.xlsx'),
            r"D:\YandexDisk\OT_online\templates\excel\Card.xlsx",  # Абсолютный путь
            os.path.join('D:', 'YandexDisk', 'OT_online', 'templates', 'excel', 'Card.xlsx'),
            # Составной абсолютный путь
        ]

        # Вывод всех возможных путей для диагностики
        for idx, path in enumerate(possible_paths):
            logger.info(f"Проверяемый путь #{idx + 1}: {path}")
            logger.info(f"Существует: {os.path.exists(path)}")

        template_path = None
        for path in possible_paths:
            if os.path.exists(path) and os.path.isfile(path):
                template_path = path
                logger.info(f"Найден шаблон по пути: {template_path}")
                break

        if template_path is None:
            paths_str = "\n".join(possible_paths)
            err_msg = f"Шаблон не найден. Проверены пути:\n{paths_str}"
            logger.error(err_msg)
            if request:
                messages.error(request, "Файл шаблона Excel не найден")
            return HttpResponseBadRequest("Файл шаблона не найден. Проверьте логи.")

        # Загружаем шаблон
        try:
            logger.info(f"Загрузка шаблона из {template_path}")
            wb = load_workbook(template_path)
            logger.info(f"Шаблон успешно загружен. Листы: {wb.sheetnames}")
        except Exception as e:
            logger.error(f"Ошибка при загрузке шаблона: {str(e)}", exc_info=True)
            if request:
                messages.error(request, f"Ошибка при загрузке шаблона: {str(e)}")
            return HttpResponseBadRequest(f"Ошибка при загрузке шаблона: {str(e)}")

        # Проверяем наличие нужного листа
        if "Лицевая сторона" not in wb.sheetnames:
            logger.error("В шаблоне отсутствует лист 'Лицевая сторона'")
            if request:
                messages.error(request, "В шаблоне отсутствует лист 'Лицевая сторона'")
            return HttpResponseBadRequest("Некорректный шаблон Excel")

        logger.info("Лист 'Лицевая сторона' найден")
        ws_front = wb["Лицевая сторона"]

        # Заполняем шапку карточки с проверкой наличия атрибутов
        logger.info("Заполнение шапки карточки")
        safe_set_cell_value(ws_front, 5, 2, getattr(employee, 'full_name_nominative', "") or "")

        # Проверка наличия поля gender
        gender = getattr(employee, 'gender', None)
        safe_set_cell_value(ws_front, 5, 5, "Мужской" if gender == "М" else "Женский")

        # Заполнение подразделения или отдела
        department_name = ""
        if hasattr(employee, 'department') and employee.department:
            department_name = employee.department.name
        elif hasattr(employee, 'subdivision') and employee.subdivision:
            department_name = employee.subdivision.name
        safe_set_cell_value(ws_front, 7, 2, department_name)

        # Заполнение должности
        safe_set_cell_value(ws_front, 9, 2, employee.position.position_name if employee.position else "")

        # Заполнение физических параметров
        safe_set_cell_value(ws_front, 6, 5,
                            str(getattr(employee, 'height', "")) if getattr(employee, 'height', None) else "")
        safe_set_cell_value(ws_front, 8, 5,
                            getattr(employee, 'clothing_size', "") if hasattr(employee, 'clothing_size') else "")
        safe_set_cell_value(ws_front, 9, 5,
                            getattr(employee, 'shoe_size', "") if hasattr(employee, 'shoe_size') else "")
        logger.info("Шапка карточки заполнена")

        # Динамически вставляем данные эталонных норм
        template_row = 16  # Строка-шаблон для норм СИЗ
        sign_row = 30  # Начало блока подписей

        try:
            logger.info(f"Вставка {norms_count} эталонных норм СИЗ")
            insert_data_with_template(
                ws=ws_front,
                data_list=list(base_norms),
                template_row=template_row,
                sign_row=sign_row,
                fill_func=fill_base_norm_row
            )
            logger.info("Данные норм СИЗ успешно вставлены")
        except Exception as e:
            logger.error(f"Ошибка при вставке данных в Excel: {str(e)}", exc_info=True)
            if request:
                messages.error(request, f"Ошибка при формировании Excel: {str(e)}")
            return HttpResponseBadRequest(f"Ошибка при формировании документа: {str(e)}")

        # Сохраняем результат во временный файл
        try:
            logger.info("Создание временного файла")
            tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            tmp_path = tmp_file.name
            logger.info(f"Временный файл создан: {tmp_path}")

            logger.info("Сохранение Excel в временный файл")
            wb.save(tmp_path)
            logger.info("Excel успешно сохранен")

            tmp_file.seek(0)

            # Формируем имя файла для скачивания
            safe_name = getattr(employee, 'full_name_nominative', '').replace(' ', '_') or f"employee_{employee_id}"
            filename = f"Card_{safe_name}.xlsx"
            logger.info(f"Имя файла для скачивания: {filename}")

            # Проверяем, что временный файл существует и доступен для чтения
            if not os.path.exists(tmp_path):
                logger.error(f"Временный файл не существует: {tmp_path}")
                if request:
                    messages.error(request, "Ошибка при создании временного файла")
                return HttpResponseBadRequest("Ошибка при создании временного файла")

            # Возвращаем файл
            logger.info("Подготовка FileResponse")
            response = FileResponse(
                open(tmp_path, "rb"),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            logger.info(f"Excel-карточка для сотрудника {employee_id} успешно сгенерирована")
            return response

        except Exception as e:
            logger.error(f"Ошибка при сохранении Excel-файла: {str(e)}", exc_info=True)
            if request:
                messages.error(request, f"Ошибка при сохранении файла: {str(e)}")
            return HttpResponseBadRequest(f"Ошибка при сохранении файла: {str(e)}")

    except Exception as e:
        logger.error(f"Непредвиденная ошибка при формировании Excel: {str(e)}", exc_info=True)
        if request:
            messages.error(request, f"Произошла ошибка: {str(e)}")
        return HttpResponseBadRequest(f"Произошла ошибка: {str(e)}")

    finally:
        # Очищаем временный файл, если он был создан
        if tmp_file and os.path.exists(tmp_file.name):
            try:
                logger.info(f"Удаление временного файла: {tmp_file.name}")
                os.unlink(tmp_file.name)
                logger.info("Временный файл успешно удален")
            except Exception as e:
                logger.warning(f"Не удалось удалить временный файл {tmp_file.name}: {str(e)}")