# directory/views/quiz_import_views.py
"""Views для импорта вопросов экзамена через веб-интерфейс"""

import os
import json
import tempfile
import zipfile
from pathlib import Path
from django.shortcuts import render, redirect
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from django.core.files.base import ContentFile
import openpyxl
import xlrd

from directory.forms.quiz_import_form import QuizImportForm, QuizImportConfirmForm
from directory.models import QuizCategory, Question, Answer


@staff_member_required
def quiz_import_upload(request):
    """Шаг 1: Загрузка файла и предпросмотр"""

    if request.method == 'POST':
        form = QuizImportForm(request.POST, request.FILES)

        if form.is_valid():
            category = form.cleaned_data['category']
            excel_file = request.FILES.get('excel_file')
            images_zip = request.FILES.get('images_zip')
            replace_existing = form.cleaned_data['replace_existing']

            # Сохраняем файлы во временную директорию
            temp_dir = tempfile.mkdtemp()

            # Сохраняем Excel файл (если есть)
            excel_path = None
            if excel_file:
                ext = '.xlsx' if excel_file.name.endswith('.xlsx') else '.xls'
                excel_path = os.path.join(temp_dir, f'questions{ext}')
                with open(excel_path, 'wb') as f:
                    for chunk in excel_file.chunks():
                        f.write(chunk)

            # Сохраняем ZIP с изображениями (если есть)
            images_dir = None
            if images_zip:
                images_zip_path = os.path.join(temp_dir, 'images.zip')
                with open(images_zip_path, 'wb') as f:
                    for chunk in images_zip.chunks():
                        f.write(chunk)

                # Распаковываем
                images_dir = os.path.join(temp_dir, 'images')
                os.makedirs(images_dir, exist_ok=True)

                with zipfile.ZipFile(images_zip_path, 'r') as zip_ref:
                    zip_ref.extractall(images_dir)

            # Парсим вопросы (если есть Excel файл)
            try:
                questions = []
                validation_errors = []

                if excel_path:
                    # Импорт вопросов из Excel
                    if excel_path.endswith('.xlsx'):
                        questions = parse_questions_from_xlsx(excel_path)
                    else:
                        questions = parse_questions_from_xls(excel_path)
                    validation_errors = validate_questions(questions)

                # Подготавливаем данные для сессии
                import_data = {
                    'category_id': category.id,
                    'excel_path': excel_path,
                    'images_dir': images_dir,
                    'replace_existing': replace_existing,
                    'questions': questions,
                    'validation_errors': validation_errors,
                    'images_only': not excel_path,  # Флаг: только изображения
                }

                # Сохраняем в сессию
                request.session['quiz_import_data'] = import_data

                # Переходим к предпросмотру
                return redirect('directory:quiz:quiz_import_preview')

            except Exception as e:
                messages.error(request, f'Ошибка при парсинге файла: {e}')
                # Очищаем временные файлы
                import shutil
                shutil.rmtree(temp_dir, ignore_errors=True)

    else:
        form = QuizImportForm()

    context = {
        'form': form,
        'title': 'Импорт вопросов экзамена',
    }

    return render(request, 'directory/quiz/import_upload.html', context)


@staff_member_required
def quiz_import_preview(request):
    """Шаг 2: Предпросмотр и подтверждение"""

    import_data = request.session.get('quiz_import_data')

    if not import_data:
        messages.error(request, 'Данные импорта не найдены. Загрузите файл заново.')
        return redirect('directory:quiz:quiz_import_upload')

    if request.method == 'POST':
        # Подтверждение импорта
        return quiz_import_confirm(request)

    # Получаем данные
    category = QuizCategory.objects.get(id=import_data['category_id'])
    questions = import_data['questions']
    validation_errors = import_data['validation_errors']
    replace_existing = import_data['replace_existing']
    images_only = import_data.get('images_only', False)

    # Статистика
    total_questions = len(questions)
    questions_with_correct = sum(1 for q in questions if any(a['is_correct'] for a in q['answers']))
    questions_with_errors = len(validation_errors)

    # Проверка наличия изображений
    images_count = 0
    if import_data.get('images_dir'):
        images_dir = Path(import_data['images_dir'])
        if images_dir.exists():
            # Считаем изображения в корневой директории
            images_count = len(
                list(images_dir.glob('*.jpg')) +
                list(images_dir.glob('*.jpeg')) +
                list(images_dir.glob('*.png')) +
                list(images_dir.glob('*.gif'))
            )

            # Также считаем в подпапках (на случай вложенной структуры ZIP)
            for subdir in images_dir.iterdir():
                if subdir.is_dir():
                    images_count += len(
                        list(subdir.glob('*.jpg')) +
                        list(subdir.glob('*.jpeg')) +
                        list(subdir.glob('*.png')) +
                        list(subdir.glob('*.gif'))
                    )

    # Текущее количество вопросов в разделе
    existing_questions_count = Question.objects.filter(category=category).count()

    # Итоговое количество
    if replace_existing:
        final_count = total_questions
    else:
        final_count = existing_questions_count + total_questions

    context = {
        'title': 'Предпросмотр импорта',
        'category': category,
        'questions': questions[:5],  # Показываем первые 5
        'total_questions': total_questions,
        'questions_with_correct': questions_with_correct,
        'questions_with_errors': questions_with_errors,
        'validation_errors': validation_errors[:10],  # Первые 10 ошибок
        'replace_existing': replace_existing,
        'existing_questions_count': existing_questions_count,
        'final_count': final_count,
        'images_count': images_count,
        'images_only': images_only,  # Режим только изображений
    }

    return render(request, 'directory/quiz/import_preview.html', context)


@staff_member_required
def quiz_import_confirm(request):
    """Шаг 3: Финальный импорт в базу данных"""

    import_data = request.session.get('quiz_import_data')

    if not import_data:
        messages.error(request, 'Данные импорта не найдены.')
        return redirect('directory:quiz:quiz_import_upload')

    try:
        category = QuizCategory.objects.get(id=import_data['category_id'])
        questions = import_data['questions']
        replace_existing = import_data['replace_existing']
        images_dir = import_data.get('images_dir')
        images_only = import_data.get('images_only', False)

        # РЕЖИМ 1: Только изображения (привязка к существующим вопросам)
        if images_only:
            if not images_dir:
                messages.error(request, 'Не найдена директория с изображениями')
                return redirect('directory:quiz:quiz_import_upload')

            # Получаем все вопросы раздела, отсортированные по order
            existing_questions = Question.objects.filter(category=category).order_by('order', 'id')

            attached_count = 0
            not_found_questions = []  # Список номеров вопросов без изображений
            errors = []

            # Отладочная информация
            import logging
            logger = logging.getLogger(__name__)
            logger.info(f'Режим: только изображения. Директория: {images_dir}')
            logger.info(f'Найдено вопросов в разделе: {existing_questions.count()}')

            # Получаем список всех изображений в архиве для отчета
            images_path = Path(images_dir)
            all_images = []
            for ext in ['*.jpg', '*.jpeg', '*.png', '*.gif']:
                all_images.extend(images_path.glob(ext))
            for subdir in images_path.iterdir():
                if subdir.is_dir():
                    for ext in ['*.jpg', '*.jpeg', '*.png', '*.gif']:
                        all_images.extend(subdir.glob(ext))

            used_images = set()  # Отслеживаем использованные изображения

            for idx, question in enumerate(existing_questions, 1):
                image_path = find_image_for_question(images_dir, idx)

                if image_path:
                    logger.info(f'Найдено изображение для вопроса #{idx}: {image_path}')
                    used_images.add(image_path)
                    try:
                        with open(image_path, 'rb') as img_file:
                            # Удаляем старое изображение если есть
                            if question.image:
                                question.image.delete(save=False)

                            question.image.save(
                                os.path.basename(image_path),
                                ContentFile(img_file.read()),
                                save=True
                            )
                        attached_count += 1
                    except Exception as e:
                        logger.error(f'Ошибка при привязке изображения к вопросу #{idx}: {e}')
                        errors.append(f'Ошибка при привязке изображения к вопросу #{idx}: {e}')
                else:
                    logger.debug(f'Изображение для вопроса #{idx} не найдено')
                    not_found_questions.append(idx)

            # Находим неиспользованные изображения
            unused_images = [str(img) for img in all_images if str(img) not in used_images]

            # Очищаем временные файлы
            import shutil
            temp_dir = os.path.dirname(images_dir)
            shutil.rmtree(temp_dir, ignore_errors=True)

            # Очищаем сессию
            del request.session['quiz_import_data']

            # Сообщения
            messages.success(request, f'✅ Изображений привязано к вопросам: {attached_count}')

            if not_found_questions:
                # Показываем первые 20 номеров вопросов без изображений
                questions_str = ', '.join(map(str, not_found_questions[:20]))
                if len(not_found_questions) > 20:
                    questions_str += f' ... (всего {len(not_found_questions)})'
                messages.warning(
                    request,
                    f'⚠️ Изображения не найдены для вопросов: {questions_str}'
                )

            if unused_images:
                # Показываем имена файлов неиспользованных изображений
                unused_names = [os.path.basename(img) for img in unused_images[:10]]
                unused_str = ', '.join(unused_names)
                if len(unused_images) > 10:
                    unused_str += f' ... (всего {len(unused_images)})'
                messages.info(
                    request,
                    f'ℹ️ Неиспользованные изображения в архиве: {unused_str}'
                )

            if errors:
                for error in errors[:5]:
                    messages.error(request, error)

            return redirect('admin:directory_quizcategory_change', category.id)

        # РЕЖИМ 2: Импорт вопросов из Excel (с опциональными изображениями)
        # Удаляем существующие вопросы если нужно
        if replace_existing:
            deleted_count = Question.objects.filter(category=category).delete()[0]
            messages.info(request, f'Удалено существующих вопросов: {deleted_count}')

        # Импортируем вопросы
        imported_count = 0
        errors = []

        for idx, question_data in enumerate(questions, 1):
            try:
                # Создаем вопрос
                question = Question.objects.create(
                    category=category,
                    question_text=question_data['text'],
                    explanation='',  # Поле оставляем пустым
                    order=idx
                )

                # Добавляем изображение если есть
                if images_dir:
                    image_path = find_image_for_question(images_dir, idx)
                    if image_path:
                        with open(image_path, 'rb') as img_file:
                            question.image.save(
                                os.path.basename(image_path),
                                ContentFile(img_file.read()),
                                save=True
                            )

                # Создаем варианты ответов
                for answer_data in question_data['answers']:
                    Answer.objects.create(
                        question=question,
                        answer_text=answer_data['text'],
                        is_correct=answer_data['is_correct'],
                        order=answer_data['order']
                    )

                imported_count += 1

            except Exception as e:
                errors.append(f'Ошибка при импорте вопроса #{idx}: {e}')

        # Очищаем временные файлы
        import shutil
        temp_path = import_data.get('excel_path')
        if temp_path:
            temp_dir = os.path.dirname(temp_path)
            shutil.rmtree(temp_dir, ignore_errors=True)

        # Очищаем сессию
        del request.session['quiz_import_data']

        # Сообщения
        messages.success(request, f'✅ Успешно импортировано вопросов: {imported_count}')

        if errors:
            for error in errors[:5]:
                messages.error(request, error)

        return redirect('admin:directory_quizcategory_change', category.id)

    except Exception as e:
        messages.error(request, f'Ошибка при импорте: {e}')
        return redirect('directory:quiz:quiz_import_upload')


@staff_member_required
def quiz_import_cancel(request):
    """Отмена импорта и очистка временных файлов"""

    import_data = request.session.get('quiz_import_data')

    if import_data:
        # Очищаем временные файлы
        import shutil
        temp_path = import_data.get('excel_path')
        if temp_path:
            temp_dir = os.path.dirname(temp_path)
            shutil.rmtree(temp_dir, ignore_errors=True)

        # Очищаем сессию
        del request.session['quiz_import_data']

    messages.info(request, 'Импорт отменен')
    return redirect('directory:quiz:quiz_import_upload')


# === ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ===


def validate_questions(questions):
    """Валидация распознанных вопросов"""
    errors = []

    for idx, question in enumerate(questions, 1):
        # Проверка: есть ли текст вопроса
        if not question.get('text'):
            errors.append(f'Вопрос #{idx}: отсутствует текст вопроса')

        # Проверка: минимум 2 варианта ответа
        if len(question.get('answers', [])) < 2:
            errors.append(f'Вопрос #{idx}: меньше 2 вариантов ответа')

        # Проверка: есть ли правильный ответ
        correct_answers = [a for a in question.get('answers', []) if a['is_correct']]

        if len(correct_answers) == 0:
            errors.append(f'Вопрос #{idx}: нет правильного ответа (жирного текста)')

        # ВАЖНО: Проверка что правильный ответ ТОЛЬКО ОДИН
        if len(correct_answers) > 1:
            errors.append(f'Вопрос #{idx}: несколько правильных ответов ({len(correct_answers)}), должен быть только ОДИН')

    return errors


def find_image_for_question(images_dir, question_number):
    """
    Поиск изображения для вопроса по номеру.
    Формат: 1.jpg, 01.jpg, 001.jpg, 2.png, 02.png и т.д.
    Также ищет в подпапках.
    """
    images_path = Path(images_dir)

    # Варианты форматирования номера: 1, 01, 001
    number_formats = [
        str(question_number),              # 1
        f"{question_number:02d}",          # 01
        f"{question_number:03d}",          # 001
    ]

    # Расширения файлов
    extensions = ['.jpg', '.jpeg', '.png', '.gif']

    # Проверяем в корневой директории
    for num_format in number_formats:
        for ext in extensions:
            image_file = images_path / f"{num_format}{ext}"
            if image_file.exists():
                return str(image_file)

    # Проверяем в подпапках (на случай, если ZIP распаковался с вложенной структурой)
    for subdir in images_path.iterdir():
        if subdir.is_dir():
            for num_format in number_formats:
                for ext in extensions:
                    image_file = subdir / f"{num_format}{ext}"
                    if image_file.exists():
                        return str(image_file)

    return None


def parse_questions_from_xlsx(filepath):
    """
    Парсинг вопросов из .xlsx файла
    Структура: Колонка B - вопрос, колонки C-F - варианты ответов
    Правильный ответ выделен жирным шрифтом
    """
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook.active

    questions = []

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=1):
        # Пропускаем пустые строки
        if not row[1].value:  # Если нет текста вопроса
            continue

        # Колонка B (индекс 1) - вопрос
        question_text = str(row[1].value).strip()

        # Колонки C, D, E, F (индексы 2-5) - варианты ответов
        answers = []
        for i in range(2, 6):  # Колонки C-F
            if i < len(row) and row[i].value:
                answer_text = str(row[i].value).strip()

                # Проверяем, жирный ли текст
                is_bold = False
                if row[i].font and row[i].font.bold:
                    is_bold = True

                answers.append({
                    'text': answer_text,
                    'is_correct': is_bold,
                    'order': i - 1  # 1, 2, 3, 4
                })

        if answers:
            questions.append({
                'text': question_text,
                'answers': answers
            })

    return questions


def parse_questions_from_xls(filepath):
    """
    Парсинг вопросов из .xls файла (старый формат Excel)
    Структура: Колонка B - вопрос, колонки C-F - варианты ответов
    Правильный ответ выделен жирным шрифтом
    """
    workbook = xlrd.open_workbook(filepath, formatting_info=True)
    sheet = workbook.sheet_by_index(0)

    questions = []

    for row_idx in range(1, sheet.nrows):  # Пропускаем заголовок
        # Колонка B (индекс 1) - вопрос
        question_cell = sheet.cell(row_idx, 1)
        if not question_cell.value:
            continue

        question_text = str(question_cell.value).strip()

        # Колонки C, D, E, F (индексы 2-5) - варианты ответов
        answers = []
        for col_idx in range(2, 6):
            if col_idx < sheet.ncols:
                answer_cell = sheet.cell(row_idx, col_idx)
                if answer_cell.value:
                    answer_text = str(answer_cell.value).strip()

                    # Проверяем, жирный ли текст в .xls
                    is_bold = False
                    try:
                        xf_index = answer_cell.xf_index
                        xf = workbook.format_map[xf_index]
                        font = workbook.font_list[xf.font_index]
                        is_bold = font.weight >= 700  # 700 = bold
                    except:
                        is_bold = False

                    answers.append({
                        'text': answer_text,
                        'is_correct': is_bold,
                        'order': col_idx - 1
                    })

        if answers:
            questions.append({
                'text': question_text,
                'answers': answers
            })

    return questions
