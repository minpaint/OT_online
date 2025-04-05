```python
# D:\YandexDisk\OT_online\apply_code_changes.py
import os
import sys

# --- КОНФИГУРАЦИЯ ---
# УКАЖИТЕ АБСОЛЮТНЫЙ ПУТЬ К КОРНЕВОЙ ПАПКЕ ВАШЕГО ПРОЕКТА
# ВАЖНО: Убедитесь, что путь указан правильно, иначе файлы могут быть перезаписаны не там!
PROJECT_ROOT = r"D:\YandexDisk\OT_online"

# --- ДАННЫЕ ДЛЯ ОБНОВЛЕНИЯ ФАЙЛОВ ---
# Ключ: Относительный путь к файлу от PROJECT_ROOT
# Значение: Новое содержимое файла
UPDATED_FILES = {
    r"directory\models\document_template.py": r"""
# D:\YandexDisk\OT_online\directory\models\document_template.py
from django.db import models
from django.core.files.storage import FileSystemStorage
from django.utils.translation import gettext_lazy as _
import logging # Добавили логгер

logger = logging.getLogger(__name__)

# Хранилище файлов для шаблонов документов
document_storage = FileSystemStorage(location='media/document_templates/')

class DocumentTemplate(models.Model):
    """
    📃 Модель для хранения шаблонов документов (DOCX файлы)

    Хранит информацию о шаблонах документов, которые используются
    для генерации документов на основе данных сотрудников.
    """

    # Типы документов - ДОБАВЛЕНА Карточка СИЗ
    DOCUMENT_TYPES = (
        ('internship_order', '🔖 Распоряжение о стажировке'),
        ('admission_order', '🔖 Распоряжение о допуске к самостоятельной работе'),
        ('knowledge_protocol', '📋 Протокол проверки знаний по охране труда'),
        ('doc_familiarization', '📝 Лист ознакомления с документами'),
        ('ppe_card', '📊 Карточка учета СИЗ (Excel)'), # <--- ДОБАВЛЕНО
    )

    name = models.CharField(_("Название шаблона"), max_length=255)
    description = models.TextField(_("Описание"), blank=True)
    document_type = models.CharField(
        _("Тип документа"),
        max_length=50,
        choices=DOCUMENT_TYPES
    )
    # Для Карточки СИЗ это поле может быть не обязательным, т.к. шаблон Excel может быть фиксированным
    template_file = models.FileField(
        _("Файл шаблона (DOCX)"),
        upload_to='document_templates/',
        storage=document_storage,
        blank=True, # <--- Сделали необязательным
        null=True, # <--- Сделали необязательным
        help_text=_("Обязательно для DOCX шаблонов. Для Excel Карточки СИЗ не используется.")
    )
    is_active = models.BooleanField(_("Активен"), default=True)
    created_at = models.DateTimeField(_("Дата создания"), auto_now_add=True)
    updated_at = models.DateTimeField(_("Дата обновления"), auto_now=True)

    def save(self, *args, **kwargs):
        # Проверка для типов документов, требующих шаблон
        if self.document_type != 'ppe_card' and not self.template_file:
             logger.warning(f"Попытка сохранить шаблон '{self.name}' типа '{self.document_type}' без файла шаблона.")
             # Можно добавить ValidationError, если это критично
             # raise ValidationError({'template_file': _('Для этого типа документа требуется файл шаблона.')})
        super().save(*args, **kwargs)


    class Meta:
        verbose_name = _("Шаблон документа")
        verbose_name_plural = _("Шаблоны документов")
        ordering = ['-updated_at']

    def __str__(self):
        return f"{self.name} ({self.get_document_type_display()})"


class GeneratedDocument(models.Model):
    """
    📄 Модель для хранения сгенерированных документов

    Хранит информацию о документах, сгенерированных на основе шаблонов.
    """
    template = models.ForeignKey(
        DocumentTemplate,
        verbose_name=_("Шаблон"),
        on_delete=models.SET_NULL, # При удалении шаблона сохраняем документ
        null=True # Может быть null, если шаблон удален
    )
    # Поле для хранения как DOCX, так и XLSX файлов
    document_file = models.FileField(
        _("Файл документа"),
        upload_to='generated_documents/%Y/%m/%d/'
    )
    employee = models.ForeignKey(
        'directory.Employee',
        verbose_name=_("Сотрудник"),
        on_delete=models.CASCADE,
        related_name="generated_documents" # Изменено для ясности
    )
    created_by = models.ForeignKey(
        'auth.User',
        verbose_name=_("Создан пользователем"),
        on_delete=models.SET_NULL,
        null=True,
        blank=True
    )
    created_at = models.DateTimeField(_("Дата создания"), auto_now_add=True)
    # Храним контекст, использованный для генерации
    document_data = models.JSONField(
        _("Данные документа"),
        default=dict,
        blank=True,
        help_text=_("Данные, использованные для генерации документа")
    )

    class Meta:
        verbose_name = _("Сгенерированный документ")
        verbose_name_plural = _("Сгенерированные документы")
        ordering = ['-created_at']

    def __str__(self):
        # Отображаем тип документа из связанного шаблона, если он есть
        doc_type_display = self.template.get_document_type_display() if self.template else _("Документ")
        return f"{doc_type_display} для {self.employee} ({self.created_at.strftime('%d.%m.%Y')})"
""",
    r"directory\forms\document_forms.py": r"""
# D:\YandexDisk\OT_online\directory\forms\document_forms.py
"""
📝 Формы для работы с документами

Этот модуль содержит формы для выбора и настройки параметров документов.
"""
from django import forms
from django.utils.translation import gettext_lazy as _
from django.forms import widgets
from crispy_forms.helper import FormHelper
from crispy_forms.layout import Layout, Fieldset, ButtonHolder, Submit, Div, HTML, Button, Field

from directory.models.document_template import DocumentTemplate # <--- Используется для DOCUMENT_TYPES
from directory.models import Employee


class DocumentSelectionForm(forms.Form):
    """
    Форма для выбора типов документов для генерации
    """
    # Исключаем тип 'ppe_card' из стандартного выбора CheckboxSelectMultiple,
    # так как его генерация будет обрабатываться отдельно.
    DOC_CHOICES = [choice for choice in DocumentTemplate.DOCUMENT_TYPES if choice[0] != 'ppe_card']

    document_types = forms.MultipleChoiceField(
        label=_("Типы документов (DOCX)"), # Уточнили в метке
        choices=DOC_CHOICES, # Используем отфильтрованные choices
        widget=forms.CheckboxSelectMultiple,
        required=False, # Делаем необязательным, так как Карточка СИЗ выбирается отдельно
        help_text=_("Выберите документы DOCX для генерации"), # Изменили help_text
        # Устанавливаем типы по умолчанию (кроме Карточки СИЗ)
        # initial=[choice[0] for choice in DOC_CHOICES] # Убрали initial по умолчанию
    )

    # Добавляем отдельный чекбокс для Карточки СИЗ
    generate_ppe_card = forms.BooleanField(
        label=_("📊 Сгенерировать Карточку учета СИЗ (Excel)"),
        required=False,
        initial=False, # По умолчанию НЕ включен
        widget=forms.CheckboxInput(attrs={'class': 'form-check-input'})
    )

    employee_id = forms.IntegerField(
        widget=forms.HiddenInput(),
        required=True
    )

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.helper = FormHelper()
        self.helper.form_method = 'post'
        self.helper.form_id = 'document-selection-form'
        self.helper.layout = Layout(
            'employee_id',
            Fieldset(
                _('Выберите документы для генерации'),
                # Размещаем стандартные типы и Карточку СИЗ рядом
                Div(
                     Field('document_types'),
                     css_class='mb-3'
                ),
                Div( # Оборачиваем чекбокс СИЗ в form-check для правильного отображения
                    Field('generate_ppe_card', css_class='form-check-input'),
                    HTML('<label class="form-check-label" for="id_generate_ppe_card">📊 Сгенерировать Карточку учета СИЗ (Excel)</label>'),
                    css_class='form-check mb-3'
                ),
                css_class='mb-3'
            ),

            ButtonHolder(
                Submit('next', _('Далее (Предпросмотр)'), css_class='btn-primary'),
                Button('cancel', _('Отмена'), css_class='btn-secondary',
                       onclick="window.history.back();")
            )
        )

    def clean(self):
        cleaned_data = super().clean()
        document_types = cleaned_data.get('document_types')
        generate_ppe_card = cleaned_data.get('generate_ppe_card')

        if not document_types and not generate_ppe_card:
            raise forms.ValidationError(_('Необходимо выбрать хотя бы один тип документа или Карточку СИЗ.'))

        return cleaned_data

# --- Формы InternshipOrderForm и AdmissionOrderForm УДАЛЕНЫ ---
# Они больше не нужны как отдельные представления, т.к. данные собираются
# в _prepare_document_context и редактируются на странице предпросмотра.

# --- Форма DocumentPreviewForm УДАЛЕНА ---
# Она больше не нужна, так как предпросмотр теперь отображается динамически
# в шаблоне documents_preview.html без отдельной формы.
""",
    r"directory\utils\docx_generator.py": r"""
# D:\YandexDisk\OT_online\directory\utils\docx_generator.py
"""
📄 Модуль для генерации документов Word

Этот модуль содержит функции для работы с шаблонами DOCX и
генерации документов на основе данных из системы.
"""
import os
import uuid
from typing import Dict, Any, Optional
import datetime
from docxtpl import DocxTemplate
from django.conf import settings
from django.core.files.base import ContentFile
import logging # Добавили импорт logging

from directory.models.document_template import DocumentTemplate, GeneratedDocument
from directory.utils.declension import decline_full_name, decline_phrase, get_initials_from_name
# Импортируем утилиты для получения данных
from directory.views.documents.utils import (
    get_internship_leader_position, get_internship_leader_name,
    get_internship_leader_initials, get_director_info, MISSING_DATA_PLACEHOLDER
)

# Настройка логгера
logger = logging.getLogger(__name__)

def get_template_path(template_id: int) -> Optional[str]:
    """
    Получает полный путь к файлу шаблона DOCX.
    Args:
        template_id (int): ID шаблона документа
    Returns:
        str: Полный путь к файлу шаблона или None, если файл не найден или шаблон не имеет файла.
    Raises:
        DocumentTemplate.DoesNotExist: Если шаблон с таким ID не найден.
    """
    template = DocumentTemplate.objects.get(id=template_id)
    if template.template_file:
        # Используем безопасное соединение путей
        path = os.path.join(settings.MEDIA_ROOT, str(template.template_file.name))
        # Проверяем существование файла
        if os.path.exists(path) and os.path.isfile(path):
            return path
        else:
            logger.warning(f"Файл шаблона DOCX для DocumentTemplate ID {template_id} не найден по пути: {path}")
            return None
    else:
        logger.warning(f"DocumentTemplate ID {template_id} (тип: {template.document_type}) не имеет связанного файла шаблона DOCX.")
        return None


def prepare_employee_context(employee) -> Dict[str, Any]:
    """
    Подготавливает базовый контекст с данными сотрудника для шаблона документа.
    Не включает данные, зависящие от типа документа (например, подписантов).
    Args:
        employee: Объект модели Employee
    Returns:
        Dict[str, Any]: Словарь с базовыми данными для заполнения шаблона
    """
    now = datetime.datetime.now()
    context = {
        # Данные сотрудника
        'fio_nominative': employee.full_name_nominative,
        'fio_genitive': decline_full_name(employee.full_name_nominative, 'gent'),
        'fio_dative': decline_full_name(employee.full_name_nominative, 'datv'),
        'fio_accusative': decline_full_name(employee.full_name_nominative, 'accs'),
        'fio_instrumental': decline_full_name(employee.full_name_nominative, 'ablt'),
        'fio_prepositional': decline_full_name(employee.full_name_nominative, 'loct'),
        'fio_initials': get_initials_from_name(employee.full_name_nominative),
        # Должность
        'position_nominative': employee.position.position_name if employee.position else "",
        'position_genitive': decline_phrase(employee.position.position_name, 'gent') if employee.position else "",
        'position_dative': decline_phrase(employee.position.position_name, 'datv') if employee.position else "",
        'position_accusative': decline_phrase(employee.position.position_name, 'accs') if employee.position else "",
        'position_instrumental': decline_phrase(employee.position.position_name, 'ablt') if employee.position else "",
        'position_prepositional': decline_phrase(employee.position.position_name, 'loct') if employee.position else "",
        # Структура
        'department': employee.department.name if employee.department else "",
        'department_genitive': decline_phrase(employee.department.name, 'gent') if employee.department else "",
        'department_dative': decline_phrase(employee.department.name, 'datv') if employee.department else "",
        'subdivision': employee.subdivision.name if employee.subdivision else "",
        'subdivision_genitive': decline_phrase(employee.subdivision.name, 'gent') if employee.subdivision else "",
        'subdivision_dative': decline_phrase(employee.subdivision.name, 'datv') if employee.subdivision else "",
        # Организация
        'organization_name': employee.organization.short_name_ru if employee.organization else "",
        'organization_full_name': employee.organization.full_name_ru if employee.organization else "",
        # Даты
        'current_date': now.strftime("%d.%m.%Y"),
        'current_day': now.strftime("%d"),
        'current_month': now.strftime("%m"),
        'current_year': now.strftime("%Y"),
        'current_year_short': now.strftime("%y"),
        # Прочее
        'location': employee.organization.location if employee.organization and hasattr(employee.organization, 'location') and employee.organization.location else "г. Минск",
        # Место для номера документа (заполняется на предпросмотре)
        'order_number': "",
        # УБРАНЫ значения по умолчанию для подписантов и руководителя стажировки
    }
    return context


def generate_docx_from_template(template_id: int, context: Dict[str, Any],
                               employee, user=None) -> Optional[GeneratedDocument]:
    """
    Генерирует документ DOCX на основе шаблона и ПОЛНОГО контекста данных.
    Args:
        template_id (int): ID шаблона документа
        context (Dict[str, Any]): Словарь с ПОЛНЫМИ данными для заполнения шаблона
        employee: Объект модели Employee
        user: Пользователь, создающий документ (опционально)
    Returns:
        Optional[GeneratedDocument]: Объект сгенерированного документа или None при ошибке
    """
    try:
        template = DocumentTemplate.objects.get(id=template_id)
        template_path = get_template_path(template_id)

        if not template_path:
             logger.error(f"Файл шаблона DOCX для ID {template_id} не найден или не указан.")
             return None

        logger.info(f"Генерация DOCX: шаблон={template_path}, сотрудник={employee.id}")
        doc = DocxTemplate(template_path)

        # Перед рендерингом убираем служебные поля из контекста
        render_context = context.copy()
        render_context.pop('employee_id', None)
        render_context.pop('missing_data_list', None)
        render_context.pop('has_missing_data', None)

        # Заменяем плейсхолдеры на пустые строки для корректного рендеринга
        for key, value in render_context.items():
            if value == MISSING_DATA_PLACEHOLDER:
                render_context[key] = ""

        doc.render(render_context)

        safe_employee_name = "".join(c if c.isalnum() else "_" for c in employee.full_name_nominative)
        filename = f"{template.document_type}_{safe_employee_name}_{uuid.uuid4().hex[:8]}.docx"
        logger.info(f"Имя сгенерированного DOCX файла: {filename}")

        file_content = ContentFile(b'')
        doc.save(file_content)
        file_content.seek(0) # Перемещаем указатель в начало файла

        generated_doc = GeneratedDocument(
            template=template,
            employee=employee,
            created_by=user,
            document_data=context # Сохраняем ПОЛНЫЙ контекст (включая missing_data)
        )
        generated_doc.document_file.save(filename, file_content, save=True) # save=True сразу сохранит модель
        logger.info(f"Сгенерированный DOCX документ сохранен: {generated_doc.pk}")

        return generated_doc

    except DocumentTemplate.DoesNotExist:
        logger.error(f"Шаблон DocumentTemplate с ID {template_id} не найден.")
        return None
    except Exception as e:
        logger.exception(f"Ошибка при генерации документа DOCX из шаблона ID {template_id}: {e}")
        return None

# --- Функции generate_internship_order и generate_admission_order УДАЛЕНЫ ---
# Используется общая функция generate_document_from_template во views/documents/preview.py

def get_document_template(document_type):
    """
    Получает активный шаблон документа определенного типа (кроме 'ppe_card').
    Args:
        document_type (str): Тип документа
    Returns:
        DocumentTemplate: Объект шаблона документа или None, если шаблон не найден
    """
    try:
        # Исключаем тип 'ppe_card', так как для него нет DOCX шаблона
        if document_type == 'ppe_card':
             logger.warning("Попытка получить DOCX шаблон для Карточки СИЗ.")
             return None
        return DocumentTemplate.objects.get(document_type=document_type, is_active=True)
    except DocumentTemplate.DoesNotExist:
        logger.error(f"Активный шаблон документа типа '{document_type}' не найден")
        return None
    except DocumentTemplate.MultipleObjectsReturned:
         logger.error(f"Найдено несколько активных шаблонов для типа '{document_type}'. Возвращен первый.")
         return DocumentTemplate.objects.filter(document_type=document_type, is_active=True).first()


def generate_document_from_template(template: DocumentTemplate, employee: Employee, user=None, context: Optional[Dict[str, Any]]=None):
    """
    Генерирует документ DOCX из шаблона и контекста.
    Args:
        template: Объект модели DocumentTemplate
        employee: Объект модели Employee
        user: Пользователь, создающий документ (опционально)
        context: Словарь с ПОЛНЫМИ данными для заполнения шаблона
    Returns:
        Optional[GeneratedDocument]: Объект сгенерированного документа или None при ошибке
    """
    if not template:
        logger.error("Передан пустой шаблон для генерации документа DOCX.")
        return None
    if template.document_type == 'ppe_card':
        logger.error("Эта функция не предназначена для генерации Карточки СИЗ (Excel).")
        return None
    if not context:
        logger.error(f"Отсутствует контекст для генерации документа DOCX по шаблону ID {template.id}.")
        return None

    # Проверяем наличие флага о недостающих данных в контексте
    if context.get('has_missing_data', False):
        missing_fields = context.get('missing_data_list', [])
        logger.error(f"Попытка генерации документа '{template.name}' с недостающими данными: {missing_fields}")
        # Не генерируем документ, если есть ошибки данных
        return None

    logger.info(f"Генерация документа DOCX для {employee} по шаблону ID {template.id} (тип: {template.document_type})")
    return generate_docx_from_template(template.id, context, employee, user)

""",
    r"directory\utils\excel_export.py": r"""
# D:\YandexDisk\OT_online\directory\utils\excel_export.py
# directory/utils/excel_export.py

import os
import tempfile
import logging
from copy import copy
import uuid # Импортируем uuid

import openpyxl
from django.conf import settings
from django.http import FileResponse, HttpResponseBadRequest, HttpResponseRedirect # Добавили HttpResponseRedirect
from django.shortcuts import redirect, reverse # Добавили reverse
from django.contrib import messages
from openpyxl import load_workbook

from directory.models import Employee, SIZNorm, GeneratedDocument, DocumentTemplate # Добавили импорт GeneratedDocument, DocumentTemplate
from django.core.files.base import ContentFile # Добавили импорт ContentFile

# Настройка логгера
logger = logging.getLogger(__name__)

# --- Функции copy_row_styles, safe_set_cell_value, insert_data_with_template, fill_base_norm_row ---
# --- остаются без изменений ---
def copy_row_styles(ws, src_row, dst_row):
    try:
        max_col = ws.max_column
        for col in range(1, max_col + 1):
            src_cell = ws.cell(row=src_row, column=col)
            dst_cell = ws.cell(row=dst_row, column=col)
            if isinstance(dst_cell, openpyxl.cell.cell.MergedCell): continue
            if src_cell.has_style: dst_cell._style = copy(src_cell._style)
    except Exception as e: logger.error(f"Ошибка при копировании стилей: {str(e)}"); raise

def safe_set_cell_value(ws, row, col, value):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, openpyxl.cell.cell.MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                top_left = merged_range.min_row, merged_range.min_col
                ws.cell(row=top_left[0], column=top_left[1]).value = value
                return
        logger.warning(f"Не удалось найти главную ячейку для объединенной ячейки {cell.coordinate}")
    else: cell.value = value

def insert_data_with_template(ws, data_list, template_row, sign_row, fill_func):
    count = len(data_list)
    if count == 0: logger.info("Нет данных для вставки"); return
    if count > 1:
        logger.info(f"Вставка {count - 1} дополнительных строк перед {sign_row}")
        try:
            merged_cells_info = []
            for merged_range in list(ws.merged_cells.ranges): # Копируем список, чтобы итерировать
                 if merged_range.min_row >= sign_row:
                     merged_cells_info.append((merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col))
                     ws.unmerge_cells(str(merged_range)) # Удаляем перед вставкой
            ws.insert_rows(sign_row, amount=count - 1)
            for min_row, min_col, max_row, max_col in merged_cells_info:
                new_min_row = min_row + (count - 1); new_max_row = max_row + (count - 1)
                ws.merge_cells(start_row=new_min_row, start_column=min_col, end_row=new_max_row, end_column=max_col)
        except Exception as e: logger.error(f"Ошибка при вставке строк или объединении ячеек: {str(e)}"); raise
    for i, item in enumerate(data_list):
        row_idx = template_row + i
        if i > 0:
            try: copy_row_styles(ws, template_row, row_idx)
            except Exception as e: logger.error(f"Ошибка при копировании стилей для строки {row_idx}: {str(e)}")
        try: fill_func(ws, row_idx, item)
        except Exception as e: logger.error(f"Ошибка при заполнении строки {row_idx}: {str(e)}"); raise

def fill_base_norm_row(ws, row_idx, norm):
    siz = norm.siz; name = getattr(siz, 'name', ""); classification = getattr(siz, 'classification', "")
    unit = getattr(siz, 'unit', ""); quantity = getattr(norm, 'quantity', ""); wear_period = getattr(siz, 'wear_period', None)
    safe_set_cell_value(ws, row_idx, 1, name); safe_set_cell_value(ws, row_idx, 2, classification)
    safe_set_cell_value(ws, row_idx, 3, unit); safe_set_cell_value(ws, row_idx, 4, quantity)
    if wear_period == 0: safe_set_cell_value(ws, row_idx, 5, "До износа")
    elif wear_period is not None: safe_set_cell_value(ws, row_idx, 5, wear_period)
    else: safe_set_cell_value(ws, row_idx, 5, "")


def generate_card_excel(request, employee_id, save_to_db=False, user=None):
    """
    Генерирует Excel-файл на основе шаблона Card.xlsx.
    Может либо вернуть FileResponse, либо сохранить файл и создать запись GeneratedDocument.

    Args:
        request: Объект HttpRequest для возможности добавления сообщений.
        employee_id: Идентификатор сотрудника.
        save_to_db (bool): Если True, сохраняет файл и создает запись в БД.
        user: Пользователь, инициировавший генерацию (для save_to_db).

    Returns:
        FileResponse с файлом Excel (если save_to_db=False)
        или GeneratedDocument (если save_to_db=True)
        или HttpResponseRedirect в случае ошибки.
    """
    logger.info(f"Генерация Excel-карточки для сотрудника {employee_id}, save_to_db={save_to_db}")
    tmp_file = None # Инициализируем переменную для временного файла
    referrer_url = request.META.get('HTTP_REFERER', reverse('directory:home')) # URL для редиректа

    try:
        employee = Employee.objects.select_related('position', 'organization', 'subdivision', 'department').get(pk=employee_id)
        logger.info(f"Сотрудник найден: {employee.full_name_nominative}")

        if not employee.position:
            logger.error(f"У сотрудника {employee_id} не указана должность")
            if request: messages.error(request, "У сотрудника не указана должность. Невозможно сгенерировать Карточку СИЗ.")
            return redirect(referrer_url)

        # Проверка на обязательные поля для Карточки СИЗ
        missing_ppe_data = []
        if not employee.height: missing_ppe_data.append('Рост')
        if not employee.clothing_size: missing_ppe_data.append('Размер одежды')
        if not employee.shoe_size: missing_ppe_data.append('Размер обуви')
        # Можно добавить другие проверки

        if missing_ppe_data and save_to_db: # Проверяем только если сохраняем в БД (т.е. из preview)
             missing_str = ", ".join(missing_ppe_data)
             logger.warning(f"Недостающие данные для Карточки СИЗ сотрудника {employee_id}: {missing_str}")
             if request: messages.warning(request, f"Карточка СИЗ будет сгенерирована, но отсутствуют данные: {missing_str}.")
             # Не прерываем генерацию, но предупреждаем

        logger.info(f"Должность сотрудника: {employee.position.position_name}")

        base_norms = SIZNorm.objects.filter(position=employee.position, condition='').select_related('siz')
        norms_count = base_norms.count()
        logger.info(f"Найдено {norms_count} эталонных норм СИЗ")
        if not base_norms.exists():
            logger.warning(f"Для должности {employee.position} не найдены эталонные нормы СИЗ")
            if request: messages.warning(request, "Для данной должности не найдены эталонные нормы СИЗ. Карточка будет сгенерирована без норм.")

        # Путь к шаблону
        template_path = os.path.join(settings.BASE_DIR, 'templates', 'excel', 'Card.xlsx')
        if not os.path.exists(template_path):
            err_msg = f"Шаблон Excel не найден по пути: {template_path}"
            logger.error(err_msg)
            if request: messages.error(request, "Файл шаблона Excel Карточки СИЗ не найден.")
            return redirect(referrer_url)

        logger.info(f"Загрузка шаблона из {template_path}")
        wb = load_workbook(template_path)
        ws_front = wb["Лицевая сторона"] # Предполагаем, что лист существует

        # --- Заполнение шапки ---
        logger.info("Заполнение шапки карточки")
        safe_set_cell_value(ws_front, 5, 2, getattr(employee, 'full_name_nominative', "") or "")
        gender = "Мужской"
        try: from directory.views.siz_issued import determine_gender_from_patronymic; gender = determine_gender_from_patronymic(employee.full_name_nominative)
        except Exception as e: logger.warning(f"Ошибка при определении пола: {e}")
        safe_set_cell_value(ws_front, 5, 5, gender)
        department_name = ""
        if employee.department: department_name = employee.department.name
        elif employee.subdivision: department_name = employee.subdivision.name
        safe_set_cell_value(ws_front, 7, 2, department_name)
        safe_set_cell_value(ws_front, 9, 2, employee.position.position_name)
        safe_set_cell_value(ws_front, 6, 5, str(getattr(employee, 'height', "")) or "")
        safe_set_cell_value(ws_front, 8, 5, getattr(employee, 'clothing_size', "") or "")
        safe_set_cell_value(ws_front, 9, 5, getattr(employee, 'shoe_size', "") or "")
        logger.info("Шапка карточки заполнена")

        # --- Вставка норм ---
        template_row = 16; sign_row = 30
        if base_norms.exists():
            logger.info(f"Вставка {norms_count} эталонных норм СИЗ")
            insert_data_with_template(ws=ws_front, data_list=list(base_norms), template_row=template_row, sign_row=sign_row, fill_func=fill_base_norm_row)
            logger.info("Данные норм СИЗ успешно вставлены")

        # --- Сохранение и возврат ---
        safe_name = "".join(c if c.isalnum() else "_" for c in employee.full_name_nominative) or f"employee_{employee_id}"
        filename = f"Card_{safe_name}_{uuid.uuid4().hex[:8]}.xlsx"

        output = BytesIO() # Используем BytesIO для сохранения в память
        wb.save(output)
        output.seek(0)

        if save_to_db:
            logger.info("Сохранение Excel в БД")
            try:
                template, _ = DocumentTemplate.objects.get_or_create(document_type='ppe_card', defaults={'name': 'Карточка учета СИЗ', 'is_active': True})
                generated_doc = GeneratedDocument(template=template, employee=employee, created_by=user, document_data={'employee_id': employee_id, 'generated_type': 'excel', 'missing_ppe_data': missing_ppe_data})
                generated_doc.document_file.save(filename, ContentFile(output.read()), save=True)
                logger.info(f"Excel-карточка сохранена в БД: {generated_doc.pk}")
                return generated_doc # Возвращаем созданный объект
            except Exception as e:
                 logger.exception(f"Ошибка при сохранении Excel-карточки в БД: {e}")
                 if request: messages.error(request, f"Ошибка сохранения Excel в БД: {e}")
                 return redirect(referrer_url)
        else:
            logger.info("Подготовка FileResponse для Excel")
            response_output = BytesIO(output.read()); response_output.seek(0)
            response = FileResponse(response_output, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            logger.info(f"Excel-карточка для сотрудника {employee_id} готова к скачиванию")
            return response

    except Employee.DoesNotExist:
        logger.error(f"Сотрудник с ID {employee_id} не найден при генерации Excel")
        if request: messages.error(request, f"Сотрудник с ID {employee_id} не найден")
        return redirect(referrer_url)
    except Exception as e:
        logger.exception(f"Непредвиденная ошибка при формировании Excel: {e}")
        if request: messages.error(request, f"Произошла ошибка при создании Excel: {e}")
        return redirect(referrer_url)
""",
    r"directory\urls.py": r"""
# D:\YandexDisk\OT_online\directory\urls.py
from django.urls import path, include, reverse_lazy
from django.contrib.auth import views as auth_views
from django.shortcuts import redirect
from django.contrib.auth import logout
import logging # Импортируем logging

from .views import siz # Импорт view для SIZ
from .views import siz_issued  # Импорт модуля siz_issued

# Импортируем представления из новой модульной структуры документов
from directory.views.documents import (
    DocumentSelectionView,
    DocumentsPreviewView,
    update_document_data,
    GeneratedDocumentListView,
    GeneratedDocumentDetailView,
    document_download,
    # generate_ppe_card_view # Убираем импорт, используем из siz_issued
)

from directory.views import (
    HomePageView,
    EmployeeListView, EmployeeCreateView, EmployeeUpdateView, EmployeeDeleteView, EmployeeHiringView,
    PositionListView, PositionCreateView, PositionUpdateView, PositionDeleteView,
    UserRegistrationView,
)

from directory.autocomplete_views import (
    OrganizationAutocomplete, SubdivisionAutocomplete, DepartmentAutocomplete,
    PositionAutocomplete, DocumentAutocomplete, EquipmentAutocomplete, SIZAutocomplete,
)

app_name = 'directory'
logger = logging.getLogger(__name__)

# Представление для выхода
def logout_view(request):
    logout(request)
    return redirect('directory:auth:login')

# URL автодополнения
autocomplete_patterns = [
    path('organization/', OrganizationAutocomplete.as_view(), name='organization-autocomplete'),
    path('subdivision/', SubdivisionAutocomplete.as_view(), name='subdivision-autocomplete'),
    path('department/', DepartmentAutocomplete.as_view(), name='department-autocomplete'),
    path('position/', PositionAutocomplete.as_view(), name='position-autocomplete'),
    path('document/', DocumentAutocomplete.as_view(), name='document-autocomplete'),
    path('equipment/', EquipmentAutocomplete.as_view(), name='equipment-autocomplete'),
    path('siz/', SIZAutocomplete.as_view(), name='siz-autocomplete'),
]

# URL сотрудников
employee_patterns = [
    path('', EmployeeListView.as_view(), name='employee_list'),
    path('create/', EmployeeCreateView.as_view(), name='employee_create'),
    path('hire/', EmployeeHiringView.as_view(), name='employee_hire'),
    path('<int:pk>/update/', EmployeeUpdateView.as_view(), name='employee_update'),
    path('<int:pk>/delete/', EmployeeDeleteView.as_view(), name='employee_delete'),
]

# URL должностей
position_patterns = [
    path('', PositionListView.as_view(), name='position_list'),
    path('create/', PositionCreateView.as_view(), name='position_create'),
    path('<int:pk>/update/', PositionUpdateView.as_view(), name='position_update'),
    path('<int:pk>/delete/', PositionDeleteView.as_view(), name='position_delete'),
]

# URL документов (обновлено)
document_patterns = [
    path('', GeneratedDocumentListView.as_view(), name='document_list'),
    path('<int:pk>/', GeneratedDocumentDetailView.as_view(), name='document_detail'),
    path('<int:pk>/download/', document_download, name='document_download'),
    path('selection/<int:employee_id>/', DocumentSelectionView.as_view(), name='document_selection'),
    path('preview/', DocumentsPreviewView.as_view(), name='documents_preview'),
    path('api/update-preview-data/', update_document_data, name='update_preview_data'),
    # URL для генерации Excel Карточки СИЗ (вызывается из preview)
    # Используем представление из siz_issued
    path('generate-ppe-card/<int:employee_id>/', siz_issued.export_personal_card_excel_view, name='generate_ppe_card'),
]

# URL оборудования
equipment_patterns = []

# URL СИЗ (обновлено)
siz_patterns = [
    path('', siz.SIZListView.as_view(), name='siz_list'),
    path('norms/create/', siz.SIZNormCreateView.as_view(), name='siznorm_create'),
    path('norms/api/', siz.siz_by_position_api, name='siz_api'),
    path('personal-card/<int:employee_id>/pdf/', siz_issued.export_personal_card_pdf, name='siz_personal_card_pdf'),
    # URL для Excel Карточки СИЗ (для прямого доступа, если нужно)
    path('personal-card/<int:employee_id>/excel/', siz_issued.export_personal_card_excel_view, name='siz_personal_card_excel'),
    path('issue-selected/<int:employee_id>/', siz_issued.issue_selected_siz, name='issue_selected_siz'),
    path('issue/', siz_issued.SIZIssueFormView.as_view(), name='siz_issue'),
    path('issue/employee/<int:employee_id>/', siz_issued.SIZIssueFormView.as_view(), name='siz_issue_for_employee'),
    path('personal-card/<int:employee_id>/', siz_issued.SIZPersonalCardView.as_view(), name='siz_personal_card'),
    path('return/<int:siz_issued_id>/', siz_issued.SIZIssueReturnView.as_view(), name='siz_return'),
    # API
    path('api/positions/<int:position_id>/siz-norms/', siz.get_position_siz_norms, name='api_position_siz_norms'),
    path('api/employees/<int:employee_id>/issued-siz/', siz_issued.employee_siz_issued_list, name='api_employee_issued_siz'),
    path('api/siz/<int:siz_id>/', siz.get_siz_details, name='api_siz_details'),
]

# URL аутентификации
auth_patterns = [
    path('login/', auth_views.LoginView.as_view(template_name='registration/login.html', redirect_authenticated_user=True), name='login'),
    path('logout/', logout_view, name='logout'),
    path('register/', UserRegistrationView.as_view(), name='register'),
    path('password_reset/', auth_views.PasswordResetView.as_view(template_name='registration/password_reset_form.html', email_template_name='registration/password_reset_email.html', success_url=reverse_lazy('directory:auth:password_reset_done')), name='password_reset'),
    path('password_reset/done/', auth_views.PasswordResetDoneView.as_view(template_name='registration/password_reset_done.html'), name='password_reset_done'),
    path('reset/<uidb64>/<token>/', auth_views.PasswordResetConfirmView.as_view(template_name='registration/password_reset_confirm.html', success_url=reverse_lazy('directory:auth:password_reset_complete')), name='password_reset_confirm'),
    path('reset/done/', auth_views.PasswordResetCompleteView.as_view(template_name='registration/password_reset_complete.html'), name='password_reset_complete'),
]

# Основные URL
urlpatterns = [
    path('', HomePageView.as_view(), name='home'),
    path('auth/', include((auth_patterns, 'auth'))),
    path('autocomplete/', include(autocomplete_patterns)),
    path('employees/', include((employee_patterns, 'employees'))),
    path('positions/', include((position_patterns, 'positions'))),
    path('documents/', include((document_patterns, 'documents'))),
    path('equipment/', include((equipment_patterns, 'equipment'))),
    path('siz/', include((siz_patterns, 'siz'))),
]

logger.debug("URL-шаблоны для 'directory' успешно загружены.")
""",
    r"directory\views\__init__.py": r"""
# D:\YandexDisk\OT_online\directory\views\__init__.py
from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db.models import Prefetch, Q
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

from directory.forms import EmployeeHiringForm
from directory.models import (
    Organization,
    StructuralSubdivision,
    Department,
    Employee,
    Position
)
from .auth import UserRegistrationView

# Импортируем представления для сотрудников
from .employees import (
    EmployeeListView,
    EmployeeCreateView,
    EmployeeUpdateView,
    EmployeeDeleteView,
    EmployeeHiringView,
    get_subdivisions
)

# Импортируем представления для должностей
from .positions import (
    PositionListView,
    PositionCreateView,
    PositionUpdateView,
    PositionDeleteView,
    get_positions,
    get_departments
)

# Импортируем представления из новой модульной структуры документов
from directory.views.documents import (
    DocumentSelectionView,
    # Убраны отдельные формы
    # InternshipOrderFormView,
    # AdmissionOrderFormView,
    DocumentsPreviewView,
    GeneratedDocumentListView,
    GeneratedDocumentDetailView,
    document_download,
    update_document_data
)

# Импортируем представления для СИЗ
from . import siz # Импорт модуля siz
from .siz_issued import (
    SIZIssueFormView,
    SIZPersonalCardView,
    SIZIssueReturnView,
    employee_siz_issued_list,
    export_personal_card_excel_view, # Импорт для Excel
    export_personal_card_pdf, # Импорт для PDF
    issue_selected_siz # Импорт для массовой выдачи
)


class HomePageView(LoginRequiredMixin, TemplateView):
    """
    🏠 Главная страница с древовидным списком сотрудников
    (Логика остается без изменений, только импорты выше)
    """
    template_name = 'directory/home.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = '🏠 Главная'
        user = self.request.user
        if user.is_superuser: allowed_orgs = Organization.objects.all()
        elif hasattr(user, 'profile'): allowed_orgs = user.profile.organizations.all()
        else: allowed_orgs = Organization.objects.none()
        search_query = self.request.GET.get('search', '')
        context['search_query'] = search_query # Pass search query to template
        organizations_data = []
        filtered_employee_ids = set()
        if search_query:
             filtered_employees = Employee.objects.filter( (Q(full_name_nominative__icontains=search_query) | Q(position__position_name__icontains=search_query)) & Q(organization__in=allowed_orgs) ).select_related('position', 'organization', 'subdivision', 'department')
             filtered_employee_ids = set(filtered_employees.values_list('id', flat=True))
             org_ids_with_matches = set(filtered_employees.values_list('organization_id', flat=True))
             allowed_orgs = allowed_orgs.filter(id__in=org_ids_with_matches)
        employee_prefetch_qs = Employee.objects.select_related('position').filter(id__in=filtered_employee_ids) if search_query else Employee.objects.select_related('position')
        employee_prefetch_attr = 'filtered_employees_list' if search_query else 'all_employees_list'
        employee_prefetch = Prefetch('employees', queryset=employee_prefetch_qs, to_attr=employee_prefetch_attr)
        dept_employee_prefetch = Prefetch('employees', queryset=employee_prefetch_qs, to_attr=employee_prefetch_attr)
        department_prefetch = Prefetch('departments',queryset=Department.objects.prefetch_related(dept_employee_prefetch),to_attr='departments_with_employees')
        subdivision_prefetch = Prefetch('subdivisions',queryset=StructuralSubdivision.objects.prefetch_related(department_prefetch, employee_prefetch),to_attr='subdivisions_with_departments_and_employees')
        organizations_qs = allowed_orgs.prefetch_related(subdivision_prefetch, employee_prefetch)
        for org in organizations_qs:
            org_employees = getattr(org, employee_prefetch_attr, [])
            has_employees_in_org = bool(org_employees)
            has_employees_deeper = False
            org_data = {'id': org.id,'name': org.full_name_ru,'short_name': org.short_name_ru,'employees': list(org_employees),'subdivisions': []}
            subdivisions = getattr(org, 'subdivisions_with_departments_and_employees', [])
            for sub in subdivisions:
                 sub_employees = getattr(sub, employee_prefetch_attr, [])
                 has_employees_in_sub = bool(sub_employees)
                 sub_has_employees_deeper = False
                 sub_data = {'id': sub.id,'name': sub.name,'employees': list(sub_employees),'departments': []}
                 departments = getattr(sub, 'departments_with_employees', [])
                 for dept in departments:
                     dept_employees = getattr(dept, employee_prefetch_attr, [])
                     has_employees_in_dept = bool(dept_employees)
                     if has_employees_in_dept:
                         dept_data = {'id': dept.id,'name': dept.name,'employees': list(dept_employees)}
                         sub_data['departments'].append(dept_data); sub_has_employees_deeper = True
                 if has_employees_in_sub or sub_has_employees_deeper:
                     org_data['subdivisions'].append(sub_data); has_employees_deeper = True
            if has_employees_in_org or has_employees_deeper: organizations_data.append(org_data)
        context['organizations'] = organizations_data
        return context


# Экспортируем все представления
__all__ = [
    # Общие
    'HomePageView',
    'UserRegistrationView',
    # Сотрудники
    'EmployeeListView', 'EmployeeCreateView', 'EmployeeUpdateView', 'EmployeeDeleteView', 'EmployeeHiringView',
    'get_subdivisions',
    # Должности
    'PositionListView', 'PositionCreateView', 'PositionUpdateView', 'PositionDeleteView',
    'get_positions', 'get_departments',
    # Документы
    'DocumentSelectionView', 'DocumentsPreviewView', 'update_document_data',
    'GeneratedDocumentListView', 'GeneratedDocumentDetailView', 'document_download',
    # СИЗ
    'siz', # Модуль siz
    'SIZIssueFormView', 'SIZPersonalCardView', 'SIZIssueReturnView',
    'employee_siz_issued_list', 'export_personal_card_excel_view', 'export_personal_card_pdf',
    'issue_selected_siz',
]

""",
    r"directory\views\documents\__init__.py": r"""
# D:\YandexDisk\OT_online\directory\views\documents\__init__.py
"""
📄 Инициализация пакета представлений для работы с документами

Экспортирует все представления для работы с документами,
чтобы их можно было импортировать из directory.views.documents
"""
from .selection import DocumentSelectionView
# Убрали импорты отдельных форм
# from .forms import (
#     InternshipOrderFormView,
#     AdmissionOrderFormView
# )
from .preview import (
    DocumentsPreviewView,
    update_document_data
)
from .management import (
    GeneratedDocumentListView,
    GeneratedDocumentDetailView,
    document_download
)
# Импортируем новое представление для генерации Excel
from directory.views.siz_issued import export_personal_card_excel_view

__all__ = [
    'DocumentSelectionView',
    # Убрали отдельные формы из экспорта
    # 'InternshipOrderFormView',
    # 'AdmissionOrderFormView',
    'DocumentsPreviewView',
    'update_document_data',
    'GeneratedDocumentListView',
    'GeneratedDocumentDetailView',
    'document_download',
    'export_personal_card_excel_view', # Экспортируем представление для Excel СИЗ
]

""",
    r"directory\views\documents\forms.py": r"""
# D:\YandexDisk\OT_online\directory\views\documents\forms.py
"""
📝 Представления для форм создания документов

Этот файл БОЛЬШЕ НЕ СОДЕРЖИТ представлений для отдельных форм документов.
Логика перенесена в preview.py и selection.py.
Оставлен пустым для возможного будущего использования или может быть удален.
"""
# Содержимое файла удалено, так как InternshipOrderFormView и AdmissionOrderFormView больше не используются.
# Логика генерации теперь централизована в DocumentsPreviewView.
pass

""",
    r"directory\views\documents\preview.py": r"""
# D:\YandexDisk\OT_online\directory\views\documents\preview.py
"""
👁️ Представления для предпросмотра документов

Содержит представления для предпросмотра и редактирования документов перед генерацией.
"""
import json
import logging
from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.shortcuts import get_object_or_404, redirect, render # Добавили render
from django.http import JsonResponse, HttpResponseRedirect, HttpResponseBadRequest, FileResponse # Добавили FileResponse
from django.urls import reverse # Добавили reverse
from django.contrib import messages
from django.utils.translation import gettext as _
from django.views.decorators.http import require_POST
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required

from directory.models import Employee
from directory.models.document_template import DocumentTemplate, GeneratedDocument
# Убрали импорт DocumentPreviewForm
from directory.utils.docx_generator import (
    generate_document_from_template,
    get_document_template
)
# Импортируем функцию генерации Excel
from directory.utils.excel_export import generate_card_excel
# Импортируем плейсхолдер
from .utils import MISSING_DATA_PLACEHOLDER

# Настройка логгера
logger = logging.getLogger(__name__)

class DocumentsPreviewView(LoginRequiredMixin, TemplateView):
    """
    Представление для предпросмотра всех выбранных документов перед генерацией
    """
    template_name = 'directory/documents/documents_preview.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Предпросмотр документов')

        # Получаем данные для предпросмотра DOCX из сессии
        preview_data_json = self.request.session.get('preview_data')
        # Получаем флаг генерации Карточки СИЗ из сессии
        generate_ppe_card = self.request.session.get('generate_ppe_card', False)
        # Получаем ID сотрудника из сессии
        employee_id = self.request.session.get('employee_id_for_preview')

        # Проверяем, есть ли хоть что-то для отображения/генерации
        if not preview_data_json and not generate_ppe_card:
            messages.error(self.request, _('Нет данных для предпросмотра или генерации.'))
            context['error_message'] = _('Нет данных для отображения.')
            return context

        # Обработка данных для предпросмотра DOCX
        preview_data = []
        if preview_data_json:
            try:
                preview_data = json.loads(preview_data_json)
                context['preview_data'] = preview_data
            except json.JSONDecodeError:
                messages.error(self.request, _('Ошибка при чтении данных предпросмотра DOCX.'))
                context['preview_data'] = [] # Пустой список для шаблона

        context['generate_ppe_card'] = generate_ppe_card
        context['has_docx_documents'] = bool(preview_data)

        # Проверка на недостающие данные в DOCX
        docx_missing_data_fields = []
        for doc_data in preview_data:
            if doc_data.get('document_data', {}).get('has_missing_data', False):
                missing = doc_data['document_data'].get('missing_data_list', [])
                doc_type_display = dict(DocumentTemplate.DOCUMENT_TYPES).get(doc_data.get('document_type'), doc_data.get('document_type'))
                for field_desc in missing:
                    docx_missing_data_fields.append(f"'{doc_type_display}': {field_desc}")
        context['docx_missing_data_fields'] = docx_missing_data_fields
        context['docx_has_missing_data'] = bool(docx_missing_data_fields)
        if context['docx_has_missing_data']:
            messages.warning(self.request, _("Внимание! Для некоторых документов DOCX отсутствуют данные. Проверьте поля перед генерацией."))

        # Получение информации о сотруднике (нужна и для DOCX, и для Excel)
        employee = None
        if employee_id:
            try:
                employee = get_object_or_404(Employee, id=employee_id)
                context['employee'] = employee
                context['employee_id'] = employee_id # Передаем ID явно
            except Exception as e:
                logger.error(f"Ошибка получения сотрудника ID={employee_id}: {e}")
                messages.error(self.request, _('Не удалось получить информацию о сотруднике.'))
                context['error_message'] = _('Ошибка получения данных сотрудника.')
                context['preview_data'] = []
                context['generate_ppe_card'] = False
                return context
        else:
             messages.error(self.request, _('Не найден ID сотрудника.'))
             context['error_message'] = _('Отсутствует ID сотрудника.')
             context['preview_data'] = []
             context['generate_ppe_card'] = False
             return context

        # Проверка данных для Карточки СИЗ (если выбрана)
        ppe_card_missing_data_fields = []
        if generate_ppe_card and employee:
            if not employee.height: ppe_card_missing_data_fields.append('Рост')
            if not employee.clothing_size: ppe_card_missing_data_fields.append('Размер одежды')
            if not employee.shoe_size: ppe_card_missing_data_fields.append('Размер обуви')
            # Добавьте другие необходимые проверки для Карточки СИЗ

        context['ppe_card_missing_data_fields'] = ppe_card_missing_data_fields
        context['ppe_card_has_missing_data'] = bool(ppe_card_missing_data_fields)
        if context['ppe_card_has_missing_data']:
             messages.warning(self.request, _("Внимание! Для Карточки СИЗ отсутствуют данные о размерах сотрудника. Карточка будет сгенерирована без них."))


        # Словарь типов документов
        context['document_types_dict'] = dict(DocumentTemplate.DOCUMENT_TYPES)

        return context

    def post(self, request, *args, **kwargs):
        """
        Обработка POST-запроса для генерации документов.
        """
        action = request.POST.get('action')
        employee_id = request.session.get('employee_id_for_preview')

        if not employee_id:
            messages.error(request, _('Не найден ID сотрудника в сессии.'))
            return redirect('directory:home')

        if action == 'generate_docx':
            return self._generate_docx_documents(request, employee_id)
        elif action == 'generate_ppe_card':
            return self._generate_ppe_card_excel(request, employee_id)
        else:
            messages.error(request, _('Неизвестное действие.'))
            return redirect(request.META.get('HTTP_REFERER', 'directory:home'))


    def _generate_docx_documents(self, request, employee_id):
        """ Генерирует выбранные DOCX документы """
        logger.info(f"Запрос на генерацию DOCX документов для сотрудника {employee_id}")
        preview_data_json = request.session.get('preview_data')
        if not preview_data_json:
            messages.error(request, _('Не найдены данные для генерации DOCX документов.'))
            return redirect('directory:documents:documents_preview')

        try:
            preview_data = json.loads(preview_data_json)
            employee = get_object_or_404(Employee, id=employee_id)
        except json.JSONDecodeError:
            messages.error(request, _('Ошибка при чтении данных для генерации DOCX.'))
            return redirect('directory:documents:documents_preview')
        except Employee.DoesNotExist:
             messages.error(request, _('Сотрудник для генерации DOCX не найден.'))
             return redirect('directory:home')

        # Получаем обновленные данные из формы
        updated_data = {}
        for key, value in request.POST.items():
            if key.startswith('document_data_'):
                try:
                    _, doc_type, field_name = key.split('_', 2)
                    if doc_type not in updated_data: updated_data[doc_type] = {}
                    updated_data[doc_type][field_name] = value
                    logger.debug(f"Обновлено поле DOCX: тип={doc_type}, поле={field_name}")
                except ValueError: logger.warning(f"Некорректный ключ поля DOCX в POST: {key}")

        generated_docx = []
        errors_found = False

        for doc_data in preview_data:
            doc_type = doc_data.get('document_type')
            document_context = doc_data.get('document_data', {})

            if doc_type in updated_data:
                document_context.update(updated_data[doc_type])
                logger.info(f"Контекст для DOCX {doc_type} обновлен.")
                # Перепроверяем флаг недостающих данных после обновления
                document_context['has_missing_data'] = any(v == MISSING_DATA_PLACEHOLDER for v in document_context.values())
                document_context['missing_data_list'] = [k for k, v in document_context.items() if v == MISSING_DATA_PLACEHOLDER]


            if document_context.get('has_missing_data', False):
                missing_fields_str = ", ".join(document_context.get('missing_data_list', ['Неизвестные поля']))
                type_name = dict(DocumentTemplate.DOCUMENT_TYPES).get(doc_type, doc_type)
                messages.error(request, _(f"Невозможно сгенерировать '{type_name}': отсутствуют данные ({missing_fields_str})."))
                errors_found = True
                continue # Пропускаем генерацию

            template = get_document_template(doc_type)
            if not template:
                 messages.error(request, _(f"Не найден активный шаблон для документа типа '{doc_type}'."))
                 errors_found = True
                 continue

            try:
                generated_doc = generate_document_from_template(
                    template, employee, request.user, document_context
                )
                if generated_doc: generated_docx.append(generated_doc)
                else: errors_found = True; messages.error(request, _(f"Ошибка при генерации '{template.name}'."))
            except Exception as e:
                logger.exception(f"Ошибка генерации DOCX {doc_type} для {employee_id}: {e}")
                messages.error(request, _(f"Ошибка при генерации '{template.name}': {e}"))
                errors_found = True

        # Очищаем данные предпросмотра DOCX из сессии
        if 'preview_data' in request.session:
            del request.session['preview_data']; request.session.modified = True

        # Обработка результата
        if generated_docx and not errors_found:
            messages.success(request, _('DOCX документы успешно сгенерированы: {} шт.').format(len(generated_docx)))
            return redirect('directory:documents:document_list') # Всегда идем в список
        elif errors_found:
             if generated_docx: messages.warning(request, _('Некоторые DOCX документы были сгенерированы, но возникли ошибки при генерации других.'))
             else: messages.error(request, _('Не удалось сгенерировать DOCX документы из-за ошибок или отсутствия данных.'))
             # Возвращаемся на предпросмотр, если были ошибки
             # Но сначала сохраним флаг генерации СИЗ, если он был
             generate_ppe_card_flag = request.session.get('generate_ppe_card', False)
             response = redirect('directory:documents:documents_preview')
             request.session['generate_ppe_card'] = generate_ppe_card_flag # Восстанавливаем флаг
             request.session['employee_id_for_preview'] = employee_id # Восстанавливаем ID
             request.session.modified = True
             return response
        else: # Не должно произойти, если preview_data был не пуст, но на всякий случай
             messages.error(request, _('Не удалось сгенерировать DOCX документы.'))
             return redirect('directory:documents:documents_preview')


    def _generate_ppe_card_excel(self, request, employee_id):
        """ Генерирует Карточку СИЗ в формате Excel """
        logger.info(f"Запрос на генерацию Карточки СИЗ (Excel) для сотрудника {employee_id}")

        try:
            # Вызываем функцию генерации Excel, СОХРАНЯЯ в БД
            generated_doc = generate_card_excel(request, employee_id, save_to_db=True, user=request.user)

            # Очищаем флаг генерации СИЗ из сессии
            if 'generate_ppe_card' in request.session:
                 del request.session['generate_ppe_card']; request.session.modified = True

            if isinstance(generated_doc, GeneratedDocument):
                messages.success(request, _('Карточка СИЗ (Excel) успешно сгенерирована и сохранена.'))
                # Редирект на список документов, где будет и Excel файл
                return redirect('directory:documents:document_list')
            else:
                # generate_card_excel вернет редирект в случае ошибки, сообщение уже добавлено там
                logger.warning(f"Функция generate_card_excel не вернула объект GeneratedDocument для сотрудника {employee_id}")
                # В случае ошибки (например, редиректа из generate_card_excel), возвращаем этот ответ
                return generated_doc

        except Exception as e:
            logger.exception(f"Ошибка при вызове генерации Excel для сотрудника {employee_id}: {e}")
            messages.error(request, _(f'Критическая ошибка при генерации Карточки СИЗ: {e}'))
            return redirect('directory:documents:documents_preview') # Возврат на предпросмотр


@method_decorator(login_required, name='dispatch')
@require_POST
def update_document_data(request):
    """
    Обработчик AJAX-запроса для обновления данных документа в сессии
    """
    try:
        doc_type = request.POST.get('doc_type')
        field_name = request.POST.get('field_name')
        field_value = request.POST.get('field_value')

        preview_data_json = request.session.get('preview_data')
        if not preview_data_json:
            return JsonResponse({'success': False, 'error': 'Не найдены данные для предпросмотра'})

        preview_data = json.loads(preview_data_json)

        data_updated = False
        updated_doc_info = {}
        for doc_data in preview_data:
            if doc_data.get('document_type') == doc_type:
                if 'document_data' in doc_data:
                     doc_data['document_data'][field_name] = field_value
                     data_updated = True
                     # После ручного редактирования убираем поле из списка недостающих
                     if 'missing_data_list' in doc_data['document_data']:
                         # Используем list comprehension для удаления, чтобы избежать ошибок при изменении списка во время итерации
                         doc_data['document_data']['missing_data_list'] = [
                             item for item in doc_data['document_data']['missing_data_list']
                             if item != field_name # Сравниваем с описанием ошибки (может быть сложнее, чем просто имя поля)
                         ]
                         # Пересчитываем флаг has_missing_data
                         doc_data['document_data']['has_missing_data'] = bool(doc_data['document_data']['missing_data_list'])
                     updated_doc_info = doc_data # Сохраняем обновленные данные
                     break
        if not data_updated:
             logger.warning(f"Не удалось найти документ типа {doc_type} в данных сессии для обновления поля {field_name}")

        request.session['preview_data'] = json.dumps(preview_data, default=str)
        request.session.modified = True

        logger.debug(f"Данные сессии обновлены для {doc_type}, поле {field_name}")
        # Возвращаем обновленный статус недостающих данных для этого документа
        return JsonResponse({
            'success': True,
            'has_missing_data': updated_doc_info.get('document_data', {}).get('has_missing_data', False)
        })

    except Exception as e:
        logger.exception(f"Ошибка AJAX обновления данных документа: {e}")
        return JsonResponse({'success': False, 'error': str(e)})

""",
    r"directory\views\documents\selection.py": r"""
# D:\YandexDisk\OT_online\directory\views\documents\selection.py
"""
🔍 Представления для выбора типов документов

Содержит представления для выбора типов документов, которые нужно сгенерировать.
"""
import json
import logging
from django.views.generic import FormView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.shortcuts import get_object_or_404, redirect
from django.http import HttpResponseRedirect
from django.urls import reverse
from django.contrib import messages
from django.utils.translation import gettext as _

from directory.models import Employee
from directory.forms.document_forms import DocumentSelectionForm
from directory.utils.docx_generator import prepare_employee_context
from directory.views.documents.utils import ( # Импортируем обновленные утилиты
    get_internship_leader_position, get_internship_leader_name,
    get_internship_leader_initials, get_director_info,
    get_commission_members, get_safety_instructions,
    get_employee_documents, MISSING_DATA_PLACEHOLDER # <--- Импорт плейсхолдера
)

# Настройка логгера
logger = logging.getLogger(__name__)

class DocumentSelectionView(LoginRequiredMixin, FormView):
    """
    Представление для выбора типов документов для генерации
    """
    template_name = 'directory/documents/document_selection.html'
    form_class = DocumentSelectionForm

    def get_initial(self):
        initial = super().get_initial()
        employee_id = self.kwargs.get('employee_id')
        if employee_id:
            initial['employee_id'] = employee_id
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        employee_id = self.kwargs.get('employee_id')
        if employee_id:
            try:
                context['employee'] = get_object_or_404(Employee.objects.select_related(
                    'position', 'organization', 'subdivision', 'department'
                ), id=employee_id)
            except Employee.DoesNotExist:
                 messages.error(self.request, _("Сотрудник с указанным ID не найден."))
                 # Очищаем employee_id из kwargs, чтобы форма не получила некорректный ID
                 self.kwargs.pop('employee_id', None)
        context['title'] = _('Выбор типов документов')
        return context

    def form_valid(self, form):
        try:
            employee_id = form.cleaned_data.get('employee_id')
            document_types = form.cleaned_data.get('document_types', []) # Только DOCX типы
            generate_ppe_card = form.cleaned_data.get('generate_ppe_card', False) # Флаг для Excel

            # Проверка наличия сотрудника еще раз
            try:
                employee = get_object_or_404(Employee, id=employee_id)
            except Employee.DoesNotExist:
                messages.error(self.request, _("Сотрудник не найден."))
                return self.form_invalid(form)

            # Проверяем, выбран ли хоть какой-то документ
            if not document_types and not generate_ppe_card:
                messages.error(self.request, _('Необходимо выбрать хотя бы один тип документа или Карточку СИЗ.'))
                return self.form_invalid(form)

            # Подготавливаем базовый контекст
            base_context = prepare_employee_context(employee)

            # Готовим данные для предпросмотра DOCX
            preview_data_docx = []
            for doc_type in document_types:
                # Пропускаем ppe_card, если он случайно попал сюда
                if doc_type == 'ppe_card':
                    continue
                context = self._prepare_document_context(doc_type, employee, base_context)
                preview_data_docx.append({
                    'document_type': doc_type,
                    'document_data': context,
                    'employee_id': employee_id # Добавляем ID сюда
                })

            # Сохраняем данные DOCX и флаг Excel в сессию
            if preview_data_docx:
                self.request.session['preview_data'] = json.dumps(preview_data_docx, default=str)
            elif 'preview_data' in self.request.session:
                 del self.request.session['preview_data'] # Очищаем, если DOCX не выбраны

            self.request.session['generate_ppe_card'] = generate_ppe_card
            self.request.session['employee_id_for_preview'] = employee_id # Сохраняем ID для preview

            self.request.session.modified = True # Убеждаемся, что сессия сохранится

            logger.info(f"Подготовлены данные для предпросмотра: employee_id={employee_id}, docx_types={document_types}, generate_ppe_card={generate_ppe_card}")

            return redirect('directory:documents:documents_preview')

        except Exception as e:
            logger.exception(f"Ошибка при обработке формы выбора документов: {e}")
            messages.error(self.request, f"Произошла внутренняя ошибка: {str(e)}")
            return self.form_invalid(form)

    def _prepare_document_context(self, document_type, employee, base_context):
        """
        Подготавливает контекст для определенного типа документа DOCX.
        Добавляет информацию о недостающих данных.
        """
        context = base_context.copy()
        missing_data_list = [] # Список описаний недостающих полей

        # Функция для добавления данных и отслеживания недостающих
        def add_context(key, value, success, missing_message):
            context[key] = value if success else MISSING_DATA_PLACEHOLDER
            if not success:
                missing_data_list.append(missing_message)
            return success

        logger.debug(f"Подготовка контекста для DOCX типа: {document_type}, сотрудник: {employee.id}")

        if document_type == 'internship_order':
            context['order_number'] = '' # Пользователь вводит на предпросмотре
            period = getattr(employee.position, 'internship_period_days', 0) if employee.position else 0
            # Период стажировки - если 0, считаем что данные есть (не является ошибкой)
            add_context('internship_duration', period if period else 2, True, '') # Всегда success=True, но значение по умолчанию 2, если 0
            # Руководитель стажировки
            pos_ok = add_context('head_of_internship_position', *get_internship_leader_position(employee), 'Должность руководителя стажировки')
            name_ok = add_context('head_of_internship_name', *get_internship_leader_name(employee), 'ФИО руководителя стажировки')
            init_ok = add_context('head_of_internship_name_initials', *get_internship_leader_initials(employee), 'Инициалы руководителя стажировки')
            # Директор
            dir_info, dir_ok = get_director_info(employee.organization)
            add_context('director_position', dir_info['position'], dir_ok, 'Должность директора')
            add_context('director_name', dir_info['name'], dir_ok, 'ФИО директора')

        elif document_type == 'admission_order':
            context['order_number'] = ''
            # Директор
            dir_info, dir_ok = get_director_info(employee.organization)
            add_context('director_position', dir_info['position'], dir_ok, 'Должность директора')
            add_context('director_name', dir_info['name'], dir_ok, 'ФИО директора')
            # Инициалы руководителя (того же, что и для стажировки)
            leader_init, init_ok = get_internship_leader_initials(employee)
            add_context('head_of_internship_name_initials', leader_init, init_ok, 'Инициалы руководителя') # Используем то же поле, что в стажировке

        elif document_type == 'knowledge_protocol':
            context['protocol_number'] = ''
            context['knowledge_result'] = 'удовлетворительные' # По умолчанию
            # Члены комиссии
            members, members_ok = get_commission_members(employee)
            add_context('commission_members', members, members_ok, 'Члены комиссии')
            # Инструкции
            instruct, instruct_ok = get_safety_instructions(employee)
            add_context('safety_instructions', instruct, instruct_ok, 'Инструкции по ОТ')

        elif document_type == 'doc_familiarization':
            context['familiarization_date'] = base_context.get('current_date', '')
            # Документы
            docs, docs_ok = get_employee_documents(employee)
            add_context('documents_list', docs, docs_ok, 'Документы для ознакомления')

        # Добавляем информацию о недостающих данных в контекст
        context['missing_data_list'] = missing_data_list # Список недостающих полей
        context['has_missing_data'] = bool(missing_data_list) # Флаг наличия ошибок

        if context['has_missing_data']:
             logger.warning(f"Для документа {document_type} сотрудника {employee.id} отсутствуют данные: {missing_data_list}")

        return context
""",
    r"directory\views\documents\utils.py": r"""
# D:\YandexDisk\OT_online\directory\views\documents\utils.py
"""
🔧 Вспомогательные функции для работы с документами

Содержит утилиты и вспомогательные функции для работы с документами.
"""
import logging
from typing import Tuple, Optional, List, Dict, Any
from directory.utils.declension import get_initials_from_name
from directory.models import Employee, Organization, Position, Document # Добавили Document

# Настройка логгера
logger = logging.getLogger(__name__)

# Плейсхолдер для недостающих данных, который будет отображаться пользователю
MISSING_DATA_PLACEHOLDER = "НЕ НАЙДЕНО (укажите вручную)"

def _find_employee_with_flag(employee: Employee, flag_name: str, search_level: str) -> Optional[Employee]:
    """ Вспомогательная функция для поиска сотрудника с флагом в иерархии """
    filter_kwargs = {'position__' + flag_name: True}
    exclude_kwargs = {'pk': employee.pk} # Исключаем самого сотрудника

    if search_level == "department" and employee.department:
        filter_kwargs['department'] = employee.department
    elif search_level == "subdivision" and employee.subdivision:
        filter_kwargs['subdivision'] = employee.subdivision
        filter_kwargs['department__isnull'] = True # Только в подразделении
    elif search_level == "organization" and employee.organization:
        filter_kwargs['organization'] = employee.organization
        filter_kwargs['subdivision__isnull'] = True # Только в организации
        filter_kwargs['department__isnull'] = True
    else:
        logger.error(f"Некорректный уровень поиска '{search_level}' для сотрудника {employee.id}")
        return None # Некорректный уровень поиска

    # Добавляем select_related для оптимизации
    found_employee = Employee.objects.select_related('position').filter(**filter_kwargs).exclude(**exclude_kwargs).first()

    if found_employee:
         logger.info(f"Найден сотрудник '{found_employee}' с флагом '{flag_name}' на уровне '{search_level}' для {employee}")
    return found_employee

def get_employee_hierarchy_level(employee: Employee, target_employee: Optional[Employee]) -> Optional[str]:
    """ Определяет иерархический уровень target_employee относительно employee """
    if not target_employee or not employee: return None
    # Проверяем принадлежность к той же организации
    if employee.organization != target_employee.organization: return None

    # Проверяем уровни
    if employee.department and target_employee.department == employee.department: return "department"
    if employee.subdivision and target_employee.subdivision == employee.subdivision and not target_employee.department: return "subdivision"
    if not target_employee.subdivision and not target_employee.department: return "organization" # На уровне организации

    return None # Не удалось определить уровень или разные ветки

# --- Функции для получения данных руководителя стажировки ---

def get_internship_leader(employee: Employee) -> Tuple[Optional[Employee], Optional[str], bool]:
    """
    Иерархический поиск руководителя стажировки (can_be_internship_leader=True)
    Returns: tuple: (leader_obj, level, success)
    """
    for level in ["department", "subdivision", "organization"]:
        leader = _find_employee_with_flag(employee, "can_be_internship_leader", level)
        if leader:
            return leader, level, True
    logger.warning(f"Руководитель стажировки для {employee} не найден.")
    return None, None, False

def get_internship_leader_position(employee: Employee) -> Tuple[str, bool]:
    leader, _, success = get_internship_leader(employee)
    if success and leader and leader.position:
        return leader.position.position_name, True
    return MISSING_DATA_PLACEHOLDER, False

def get_internship_leader_name(employee: Employee) -> Tuple[str, bool]:
    leader, _, success = get_internship_leader(employee)
    if success and leader:
        return leader.full_name_nominative, True
    return MISSING_DATA_PLACEHOLDER, False

def get_internship_leader_initials(employee: Employee) -> Tuple[str, bool]:
    leader, _, success = get_internship_leader(employee)
    if success and leader:
        return get_initials_from_name(leader.full_name_nominative), True
    return MISSING_DATA_PLACEHOLDER, False

# --- Функции для получения данных подписанта (директора) ---

def get_document_signer(employee: Employee) -> Tuple[Optional[Employee], Optional[str], bool]:
    """
    Иерархический поиск подписанта документов (can_sign_orders=True)
    Returns: tuple: (signer_obj, level, success)
    """
    for level in ["department", "subdivision", "organization"]:
        signer = _find_employee_with_flag(employee, "can_sign_orders", level)
        if signer:
            return signer, level, True
    logger.warning(f"Подписант документов для {employee} не найден.")
    return None, None, False

def get_director_info(organization: Organization) -> Tuple[Dict[str, str], bool]:
    """ Ищет подписанта на уровне организации """
    # Создаем "фиктивного" сотрудника только для передачи в _find_employee_with_flag
    temp_employee_for_search = Employee(organization=organization)
    signer = _find_employee_with_flag(temp_employee_for_search, "can_sign_orders", "organization")

    if signer and signer.position:
        logger.info(f"Найден подписант (директор) для организации {organization}: {signer}")
        return {
            'position': signer.position.position_name,
            'name': get_initials_from_name(signer.full_name_nominative)
        }, True
    else:
        logger.warning(f"Подписант (директор) для организации {organization} не найден.")
        return {
            'position': MISSING_DATA_PLACEHOLDER,
            'name': MISSING_DATA_PLACEHOLDER
        }, False

# --- Функции для получения данных для конкретных документов ---

def get_commission_members(employee: Employee) -> Tuple[List[Dict[str, str]], bool]:
    """ Получает список членов комиссии из организации сотрудника """
    members_list = []
    any_found = False # Флаг, найден ли хоть кто-то
    org = employee.organization

    # Председатель
    chairman = Employee.objects.filter(organization=org, position__commission_role='chairman').select_related('position').first()
    chairman_data = {"role": "Председатель комиссии"}
    if chairman and chairman.position:
        chairman_data["position"] = chairman.position.position_name
        chairman_data["name"] = get_initials_from_name(chairman.full_name_nominative)
        any_found = True
    else:
        chairman_data["position"] = MISSING_DATA_PLACEHOLDER
        chairman_data["name"] = MISSING_DATA_PLACEHOLDER
    members_list.append(chairman_data)

    # Члены комиссии
    members = Employee.objects.filter(organization=org, position__commission_role='member').select_related('position')[:2]
    members_found_count = 0
    for member in members:
        member_data = {"role": "Член комиссии"}
        if member.position:
             member_data["position"] = member.position.position_name
             member_data["name"] = get_initials_from_name(member.full_name_nominative)
             any_found = True
             members_found_count += 1
        else:
             member_data["position"] = MISSING_DATA_PLACEHOLDER
             member_data["name"] = MISSING_DATA_PLACEHOLDER
        members_list.append(member_data)

    # Добиваем до 3 плейсхолдерами, если нашли меньше 2 членов
    while members_found_count < 2:
        members_list.append({"role": "Член комиссии", "position": MISSING_DATA_PLACEHOLDER, "name": MISSING_DATA_PLACEHOLDER})
        members_found_count += 1 # Считаем добавленный плейсхолдер

    success = any_found # Считаем успехом, если нашли хотя бы одного
    if not success: logger.warning(f"Ни один член комиссии не найден для организации {org}")
    return members_list, success


def get_safety_instructions(employee: Employee) -> Tuple[List[str], bool]:
    """ Получает список инструкций по ОТ из должности сотрудника """
    if employee.position and employee.position.safety_instructions_numbers:
        instructions = employee.position.safety_instructions_numbers.strip()
        if instructions:
            # Разделяем по запятой или точке с запятой, удаляем пустые строки
            instructions_list = [instr.strip() for instr in re.split(r'[;,]', instructions) if instr.strip()]
            if instructions_list:
                logger.info(f"Найдены инструкции по ОТ для {employee}: {instructions_list}")
                return instructions_list, True
    logger.warning(f"Инструкции по ОТ для {employee} не найдены или поле пустое.")
    return [MISSING_DATA_PLACEHOLDER], False


def get_employee_documents(employee: Employee) -> Tuple[List[str], bool]:
    """ Получает список документов для ознакомления из должности сотрудника """
    if employee.position:
        # Используем prefetch_related для оптимизации, если это часто вызывается
        documents = employee.position.documents.all().order_by('name')
        if documents.exists():
            documents_list = [doc.name for doc in documents]
            logger.info(f"Найдены документы для ознакомления для {employee}: {documents_list}")
            return documents_list, True
    logger.warning(f"Документы для ознакомления для {employee} не найдены.")
    return [MISSING_DATA_PLACEHOLDER], False

""",
     r"directory\views\siz_issued.py": r"""
# D:\YandexDisk\OT_online\directory\views\siz_issued.py
# 📁 directory/views/siz_issued.py
import re
import random
import logging
from django.views.generic import CreateView, DetailView, FormView, UpdateView, View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy, reverse
from django.shortcuts import get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect # Добавили HttpResponseRedirect
from django.contrib import messages
from django.utils.decorators import method_decorator
from django.views.decorators.http import require_GET
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.template.loader import get_template
from io import BytesIO
from xhtml2pdf import pisa

from directory.models import Employee, SIZIssued
from directory.forms.siz_issued import SIZIssueForm, SIZIssueMassForm, SIZIssueReturnForm
from directory.utils.pdf import render_to_pdf
# Импортируем функцию генерации Excel
from directory.utils.excel_export import generate_card_excel

logger = logging.getLogger(__name__)

# --- Функции determine_gender_from_patronymic и get_random_siz_sizes без изменений ---
def determine_gender_from_patronymic(full_name):
    name_parts = full_name.split()
    if len(name_parts) >= 3: patronymic = name_parts[2]
    else: return "Мужской"
    if re.search(r'(ич|ыч)$', patronymic, re.IGNORECASE): return "Мужской"
    elif re.search(r'(на|вна|чна)$', patronymic, re.IGNORECASE): return "Женский"
    elif re.search(r'(оглы|улы|лы)$', patronymic, re.IGNORECASE): return "Мужской"
    elif re.search(r'(кызы|зы)$', patronymic, re.IGNORECASE): return "Женский"
    return "Мужской"

def get_random_siz_sizes(gender):
    if gender == "Мужской":
        headgear = random.randint(55, 59); gloves = random.randint(15, 19) / 2; respirator = random.choice(["1", "2", "3"])
    else:
        headgear = random.randint(53, 57); gloves = random.randint(13, 17) / 2; respirator = random.choice(["1", "2", "3"])
    gas_mask = respirator
    return {'headgear': headgear, 'gloves': gloves, 'respirator': respirator, 'gas_mask': gas_mask}

# --- SIZIssueFormView без изменений ---
class SIZIssueFormView(LoginRequiredMixin, CreateView):
    model = SIZIssued; form_class = SIZIssueForm; template_name = 'directory/siz_issued/issue_form.html'
    def get_success_url(self): return reverse('directory:siz:siz_personal_card', kwargs={'employee_id': self.object.employee.id})
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs(); kwargs['user'] = self.request.user
        employee_id = self.kwargs.get('employee_id');
        if employee_id: kwargs['employee_id'] = employee_id
        return kwargs
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs); context['title'] = 'Выдача СИЗ'
        employee_id = self.kwargs.get('employee_id')
        if employee_id:
            employee = get_object_or_404(Employee, id=employee_id); context['employee'] = employee
            if employee.position:
                from directory.models.siz import SIZNorm; norms = SIZNorm.objects.filter(position=employee.position).select_related('siz')
                context['base_norms'] = norms.filter(condition='')
                condition_groups = {};
                for norm in norms.exclude(condition=''):
                    if norm.condition not in condition_groups: condition_groups[norm.condition] = []
                    condition_groups[norm.condition].append(norm)
                context['condition_groups'] = [{'name': c, 'norms': n} for c, n in condition_groups.items()]
        return context
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f"✅ СИЗ '{self.object.siz.name}' успешно выдано сотруднику {self.object.employee.full_name_nominative}")
        return response

# --- issue_selected_siz без изменений ---
@login_required
def issue_selected_siz(request, employee_id):
    if request.method == 'POST':
        employee = get_object_or_404(Employee, id=employee_id)
        selected_norm_ids = request.POST.getlist('selected_norms')
        if not selected_norm_ids: messages.warning(request, "Не выбрано ни одного СИЗ для выдачи"); return redirect('directory:siz:siz_personal_card', employee_id=employee_id)
        from directory.models.siz import SIZNorm
        norms = SIZNorm.objects.filter(id__in=selected_norm_ids).select_related('siz')
        issued_count = 0
        for norm in norms:
            existing_issued = SIZIssued.objects.filter(employee=employee, siz=norm.siz, is_returned=False).exists()
            if not existing_issued:
                SIZIssued.objects.create(employee=employee, siz=norm.siz, quantity=norm.quantity, issue_date=timezone.now().date(), condition=norm.condition, received_signature=True); issued_count += 1
        if issued_count > 0: messages.success(request, f"✅ Успешно выдано {issued_count} наименований СИЗ сотруднику {employee.full_name_nominative}")
        else: messages.info(request, "ℹ️ Ни одно СИЗ не было выдано. Возможно, выбранные СИЗ уже находятся в использовании.")
    return redirect('directory:siz:siz_personal_card', employee_id=employee_id)

# --- SIZPersonalCardView без изменений ---
class SIZPersonalCardView(LoginRequiredMixin, DetailView):
    model = Employee; template_name = 'directory/siz_issued/personal_card.html'; context_object_name = 'employee'
    def get_object(self): return get_object_or_404(Employee, id=self.kwargs.get('employee_id'))
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs); context['title'] = f'Личная карточка учета СИЗ - {self.object.full_name_nominative}'
        issued_items = SIZIssued.objects.filter(employee=self.object).select_related('siz').order_by('-issue_date'); context['issued_items'] = issued_items
        if self.object.position:
            from directory.models.siz import SIZNorm; norms = SIZNorm.objects.filter(position=self.object.position).select_related('siz')
            context['base_norms'] = norms.filter(condition=''); conditions = list(set(norm.condition for norm in norms if norm.condition)); condition_groups = []
            for condition in conditions:
                condition_norms = [norm for norm in norms if norm.condition == condition]
                if condition_norms: condition_groups.append({'name': condition, 'norms': condition_norms})
            context['condition_groups'] = condition_groups
        gender = determine_gender_from_patronymic(self.object.full_name_nominative); context['gender'] = gender
        context['siz_sizes'] = get_random_siz_sizes(gender)
        return context

# --- SIZIssueReturnView без изменений ---
class SIZIssueReturnView(LoginRequiredMixin, UpdateView):
    model = SIZIssued; form_class = SIZIssueReturnForm; template_name = 'directory/siz_issued/return_form.html'; pk_url_kwarg = 'siz_issued_id'
    def get_success_url(self): return reverse('directory:siz:siz_personal_card', kwargs={'employee_id': self.object.employee.id})
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs); context['title'] = 'Возврат СИЗ'; context['employee'] = self.object.employee
        context['siz_name'] = self.object.siz.name; context['issue_date'] = self.object.issue_date
        return context
    def form_valid(self, form):
        response = super().form_valid(form); messages.success(self.request, f"✅ СИЗ '{self.object.siz.name}' успешно возвращено")
        return response

# --- employee_siz_issued_list без изменений ---
@login_required
@require_GET
def employee_siz_issued_list(request, employee_id):
    employee = get_object_or_404(Employee, pk=employee_id)
    issued_items = SIZIssued.objects.filter(employee=employee).select_related('siz').order_by('-issue_date')
    result = {'employee_id': employee.id, 'employee_name': employee.full_name_nominative, 'position': employee.position.position_name if employee.position else "", 'organization': employee.organization.short_name_ru, 'issued_items': []}
    for item in issued_items:
        item_data = {'id': item.id, 'siz_name': item.siz.name, 'siz_classification': item.siz.classification, 'issue_date': item.issue_date.strftime('%d.%m.%Y'), 'quantity': item.quantity, 'wear_percentage': item.wear_percentage, 'is_returned': item.is_returned, 'return_date': item.return_date.strftime('%d.%m.%Y') if item.return_date else None, 'notes': item.notes, 'condition': item.condition}
        result['issued_items'].append(item_data)
    return JsonResponse(result)

# --- export_personal_card_pdf без изменений ---
def export_personal_card_pdf(request, employee_id):
    employee = get_object_or_404(Employee, pk=employee_id)
    issued_items = SIZIssued.objects.filter(employee=employee).select_related('siz').order_by('-issue_date')
    selected_norm_ids = request.GET.getlist('selected_norms')
    if not selected_norm_ids and employee.position:
        from directory.models.siz import SIZNorm
        all_norms = SIZNorm.objects.filter(position=employee.position).values_list('id', flat=True)
        selected_norm_ids = list(map(str, all_norms))
    selected_items = []
    if selected_norm_ids:
        from directory.models.siz import SIZNorm
        selected_norms = SIZNorm.objects.filter(id__in=selected_norm_ids).select_related('siz')
        for norm in selected_norms: selected_items.append({'siz': norm.siz, 'classification': norm.siz.classification, 'quantity': norm.quantity})
    base_norms = []; condition_groups = []
    if employee.position:
        from directory.models.siz import SIZNorm
        norms = SIZNorm.objects.filter(position=employee.position).select_related('siz')
        base_norms = list(norms.filter(condition=''))
        conditions = list(set(norm.condition for norm in norms if norm.condition))
        for condition in conditions:
            condition_norms = [norm for norm in norms if norm.condition == condition]
            if condition_norms: condition_groups.append({'name': condition, 'norms': condition_norms})
    context = {'employee': employee, 'issued_items': issued_items, 'base_norms': base_norms, 'condition_groups': condition_groups, 'today': timezone.now().date(), 'gender': determine_gender_from_patronymic(employee.full_name_nominative), 'siz_sizes': get_random_siz_sizes(determine_gender_from_patronymic(employee.full_name_nominative)), 'selected_items': selected_items}
    filename = f"personal_card_{employee.full_name_nominative.replace(' ', '_')}.pdf"
    template_path = 'directory/siz_issued/personal_card_pdf_landscape.html'
    try:
        pdf_options = {'page-size': 'A4', 'margin-top': '0.5cm', 'margin-right': '0.5cm', 'margin-bottom': '0.5cm', 'margin-left': '0.5cm', 'encoding': "UTF-8"}
        return render_to_pdf(template_path=template_path, context=context, filename=filename, as_attachment=True, pdf_options=pdf_options)
    except Exception as e:
        logger.error(f"Ошибка при создании PDF: {e}"); messages.error(request, f"Ошибка при создании PDF: {e}")
        return redirect('directory:siz:siz_personal_card', employee_id=employee_id)

# --- НОВОЕ ПРЕДСТАВЛЕНИЕ для генерации Excel из URL ---
@login_required
def export_personal_card_excel_view(request, employee_id):
    """
    Представление для генерации и скачивания Excel-карточки СИЗ по URL.
    Использует существующую функцию generate_card_excel.
    """
    logger.info(f"Запрос на генерацию Excel-карточки для сотрудника {employee_id} через URL")
    # Вызываем функцию генерации Excel, НЕ сохраняя в БД
    response = generate_card_excel(request, employee_id, save_to_db=False, user=request.user) # Добавили user

    # generate_card_excel вернет либо FileResponse, либо редирект с сообщением об ошибке
    return response
""",
    r"templates\base.html": r"""
{# D:\YandexDisk\OT_online\templates\base.html #}
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{% block title %}OT Online{% endblock %}</title>

    <!-- Bootstrap CSS через CDN 📡 -->
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css"
          rel="stylesheet"
          integrity="sha384-9ndCyUaIbzAi2FUVXJi0CjmCapSmO7SnpJef0486qhLnuZ2cdeRhO02iuK6FUUVM"
          crossorigin="anonymous">
    {# Font Awesome для иконок #}
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/5.15.3/css/all.min.css"
          integrity="sha512-iBBXm8fW90+nuLcSKlbmrPcLa0OT92xO1BIsZ+ywDWZCvqsWgccV3gFoRBv0z+8dLJgyAHIhR35VZc2oM/gI1w=="
          crossorigin="anonymous" referrerpolicy="no-referrer" />
    {% load static %}
    <!-- Основной стиль приложения 🎨 -->
    <link href="{% static 'directory/css/style.css' %}" rel="stylesheet">
    <!-- Select2 CSS через CDN для автодополнения 🔍 -->
    <link href="https://cdn.jsdelivr.net/npm/select2@4.1.0-rc.0/dist/css/select2.min.css" rel="stylesheet">
    {% block extra_css %}{% endblock %}
</head>
<body>
    <!-- Навигационная панель 🚀 -->
    <nav class="navbar navbar-expand-lg navbar-light bg-light sticky-top shadow-sm"> {# Добавлены sticky-top и shadow-sm #}
        <div class="container">
            <a class="navbar-brand fw-bold" href="{% url 'directory:home' %}">🏢 OT Online</a> {# Ссылка на главную #}
            <button class="navbar-toggler" type="button"
                    data-bs-toggle="collapse"
                    data-bs-target="#navbarNav"
                    aria-controls="navbarNav"
                    aria-expanded="false"
                    aria-label="Toggle navigation">
                <span class="navbar-toggler-icon"></span>
            </button>
            <div class="collapse navbar-collapse" id="navbarNav">
                <ul class="navbar-nav me-auto mb-2 mb-lg-0"> {# Добавлены mb-2 mb-lg-0 #}
                    {% if user.is_authenticated %}
                    <li class="nav-item">
                        <a class="nav-link {% if request.resolver_match.view_name == 'home' %}active{% endif %}"
                           aria-current="page" href="{% url 'directory:home' %}">
                           <i class="fas fa-home"></i> Главная
                        </a>
                    </li>
                    {# Убрали дублирование ссылок Сотрудники/Должности, т.к. они доступны с главной #}
                    {# <li class="nav-item">
                        <a class="nav-link {% if 'employees' in request.path %}active{% endif %}"
                           href="{% url 'directory:employees:employee_list' %}">👥 Сотрудники</a>
                    </li>
                    <li class="nav-item">
                        <a class="nav-link {% if 'positions' in request.path %}active{% endif %}"
                           href="{% url 'directory:positions:position_list' %}">👔 Должности</a>
                    </li> #}
                    <li class="nav-item">
                        <a class="nav-link {% if 'siz' in request.path %}active{% endif %}"
                           href="{% url 'directory:siz:siz_list' %}">
                           <i class="fas fa-hard-hat"></i> СИЗ
                        </a>
                    </li>
                     <li class="nav-item">
                        <a class="nav-link {% if 'documents' in request.path and 'preview' not in request.path %}active{% endif %}"
                           href="{% url 'directory:documents:document_list' %}">
                           <i class="fas fa-file-alt"></i> Документы
                        </a>
                    </li>
                    {% endif %}
                </ul>
                <ul class="navbar-nav">
                    {% if user.is_authenticated %}
                        {% if user.is_staff %}
                        <li class="nav-item">
                            <a class="nav-link" href="{% url 'admin:index' %}">
                                <i class="fas fa-cogs"></i> Админ
                            </a>
                        </li>
                        {% endif %}
                        <li class="nav-item dropdown">
                            <a class="nav-link dropdown-toggle" href="#"
                               id="userDropdown" role="button"
                               data-bs-toggle="dropdown" aria-expanded="false">
                               <i class="fas fa-user"></i> {{ user.get_full_name|default:user.username }}
                            </a>
                            <ul class="dropdown-menu dropdown-menu-end" aria-labelledby="userDropdown">
                                {# Можно добавить ссылку на профиль пользователя, если он будет #}
                                {# <li><a class="dropdown-item" href="#">Профиль</a></li>
                                <li><hr class="dropdown-divider"></li> #}
                                <li>
                                    <a class="dropdown-item text-danger" href="{% url 'directory:auth:logout' %}">
                                        <i class="fas fa-sign-out-alt"></i> Выйти
                                    </a>
                                </li>
                            </ul>
                        </li>
                    {% else %}
                        <li class="nav-item">
                            <a class="nav-link {% if 'login' in request.path %}active{% endif %}" href="{% url 'directory:auth:login' %}">
                                <i class="fas fa-sign-in-alt"></i> Войти
                            </a>
                        </li>
                        <li class="nav-item">
                            <a class="nav-link {% if 'register' in request.path %}active{% endif %}" href="{% url 'directory:auth:register' %}">
                                <i class="fas fa-user-plus"></i> Регистрация
                            </a>
                        </li>
                    {% endif %}
                </ul>
            </div>
        </div>
    </nav>

    <!-- Основной контент страницы 💻 -->
    <main class="container mt-4 mb-5"> {# Добавлен mb-5 #}
        {% if messages %}
            <div class="position-fixed top-0 end-0 p-3" style="z-index: 1050"> {# Позиционирование сообщений #}
            {% for message in messages %}
                <div class="toast align-items-center text-white bg-{{ message.tags|default:'info' }} border-0 show" role="alert" aria-live="assertive" aria-atomic="true" data-bs-delay="5000">
                  <div class="d-flex">
                    <div class="toast-body">
                      {% if message.tags == 'success' %} <i class="fas fa-check-circle me-2"></i>
                      {% elif message.tags == 'error' %} <i class="fas fa-exclamation-triangle me-2"></i>
                      {% elif message.tags == 'warning' %} <i class="fas fa-exclamation-circle me-2"></i>
                      {% else %} <i class="fas fa-info-circle me-2"></i> {% endif %}
                      {{ message }}
                    </div>
                    <button type="button" class="btn-close btn-close-white me-2 m-auto" data-bs-dismiss="toast" aria-label="Close"></button>
                  </div>
                </div>
            {% endfor %}
            </div>
        {% endif %}

        {% block content %}{% endblock %}
    </main>

    {# Футер можно добавить при необходимости #}
    {# <footer class="footer mt-auto py-3 bg-light border-top">
        <div class="container text-center">
            <span class="text-muted">OT Online &copy; {% now "Y" %}</span>
        </div>
    </footer> #}

    <!-- Подключаем JavaScript через CDN 📡 -->
    <script src="https://code.jquery.com/jquery-3.7.0.min.js" integrity="sha256-2Pmvv0kuTBOenSvLm6bvfBSSHrUJ+3A7x6P5Ebd07/g=" crossorigin="anonymous"></script>
    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js" integrity="sha384-geWF76RCwLtnZ8qwWowPQNguL3RmwHVBC9FhGdlKrxdiJJigb/j/68SIy3Te4Bkz" crossorigin="anonymous"></script>
    <script src="https://cdn.jsdelivr.net/npm/select2@4.1.0-rc.0/dist/js/select2.min.js"></script>
    <!-- Основной скрипт -->
    <script src="{% static 'directory/js/main.js' %}"></script>
    <script>
      // Инициализация всплывающих сообщений Bootstrap
      var toastElList = [].slice.call(document.querySelectorAll('.toast'))
      var toastList = toastElList.map(function (toastEl) {
        return new bootstrap.Toast(toastEl)
      })
    </script>

    {% block extra_js %}{% endblock %}
</body>
</html>

""",
    r"templates\directory\documents\document_selection.html": r"""
{# D:\YandexDisk\OT_online\templates\directory\documents\document_selection.html #}
{% extends 'base.html' %}
{% load static %}
{% load crispy_forms_tags %}

{% block title %}{{ title }}{% endblock %}

{% block extra_css %}
<style>
    .form-check-label {
        cursor: pointer;
    }
    .form-check {
        margin-bottom: 0.75rem; /* Отступ между чекбоксами */
    }
     .card-body {
        padding: 2rem; /* Увеличим отступы */
    }
</style>
{% endblock %}

{% block content %}
<div class="container mt-4">
    <div class="row justify-content-center">
        <div class="col-md-8 col-lg-7"> {# Немного увеличим ширину #}
            <div class="card shadow-sm">
                <div class="card-header bg-primary text-white">
                    <h4 class="mb-0"><i class="fas fa-list-ul me-2"></i>{{ title }}</h4>
                </div>
                <div class="card-body">
                    {% if employee %}
                    <div class="alert alert-light border rounded p-3 mb-4">
                        <h5 class="alert-heading"><i class="fas fa-user me-2"></i>Информация о сотруднике</h5>
                        <hr>
                        <p class="mb-1"><strong>ФИО:</strong> {{ employee.full_name_nominative }}</p>
                        <p class="mb-1"><strong>Должность:</strong> {{ employee.position.position_name|default:"-" }}</p>
                        {% if employee.department %}
                        <p class="mb-1"><strong>Отдел:</strong> {{ employee.department.name }}</p>
                        {% endif %}
                        {% if employee.subdivision %}
                        <p class="mb-1"><strong>Подразделение:</strong> {{ employee.subdivision.name }}</p>
                        {% endif %}
                        <p class="mb-0"><strong>Организация:</strong> {{ employee.organization.short_name_ru|default:"-" }}</p>
                    </div>
                    {% else %}
                     <div class="alert alert-warning">Сотрудник не найден.</div>
                    {% endif %}

                    {% if employee %} {# Показываем форму только если есть сотрудник #}
                    <form method="post" action="" id="document-selection-form">
                        {% csrf_token %}
                        {% crispy form %}
                    </form>
                    {% endif %}
                </div>
            </div>
        </div>
    </div>
</div>

{% block extra_js %}
<script>
    document.addEventListener('DOMContentLoaded', function() {
        // Проверка на наличие хотя бы одного выбранного типа документа ИЛИ Карточки СИЗ
        const form = document.getElementById('document-selection-form');
        if (form) {
            form.addEventListener('submit', function(e) {
                const docTypeCheckboxes = document.querySelectorAll('input[name="document_types"]:checked');
                const ppeCardCheckbox = document.getElementById('id_generate_ppe_card'); // Получаем чекбокс СИЗ
                const isPpeCardChecked = ppeCardCheckbox ? ppeCardCheckbox.checked : false;

                if (docTypeCheckboxes.length === 0 && !isPpeCardChecked) { // Проверяем оба условия
                    e.preventDefault();
                    // Используем Bootstrap Alert или Toast для лучшего UX
                    alert('Пожалуйста, выберите хотя бы один тип документа или Карточку СИЗ.');
                    // Можно добавить более красивое уведомление
                    // showBootstrapAlert('Пожалуйста, выберите хотя бы один тип документа или Карточку СИЗ.', 'warning');
                }
            });
        }
        // Функция для показа Bootstrap Alert (пример)
        // function showBootstrapAlert(message, type = 'danger') {
        //     const alertPlaceholder = document.getElementById('alertPlaceholder'); // Нужен div с id="alertPlaceholder" в шаблоне
        //     if (alertPlaceholder) {
        //         const wrapper = document.createElement('div');
        //         wrapper.innerHTML = [
        //             `<div class="alert alert-${type} alert-dismissible" role="alert">`,
        //             `   <div>${message}</div>`,
        //             '   <button type="button" class="btn-close" data-bs-dismiss="alert" aria-label="Close"></button>',
        //             '</div>'
        //         ].join('');
        //         alertPlaceholder.append(wrapper);
        //     }
        // }
    });
</script>
{% endblock %}

{% endblock %}

""",
    r"templates\directory\documents\documents_preview.html": r"""
{# D:\YandexDisk\OT_online\templates\directory\documents\documents_preview.html #}
{% extends 'base.html' %}
{% load static %}
{% load i18n %}
{% load document_filters %} {# Загружаем наш фильтр #}

{% block title %}{{ title }}{% endblock %}

{% block extra_css %}
<style>
    .document-preview-section {
        margin-bottom: 2rem;
        border: 1px solid #e0e0e0;
        border-radius: 8px;
        background-color: #fff;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05);
    }

    .document-preview-header {
        background-color: #f8f9fa;
        padding: 1rem 1.5rem;
        border-bottom: 1px solid #e0e0e0;
        border-top-left-radius: 8px;
        border-top-right-radius: 8px;
        display: flex;
        justify-content: space-between;
        align-items: center;
    }
    .document-preview-header h5 {
        margin-bottom: 0;
        font-weight: 600;
        color: #495057;
    }

    .document-preview-body {
        padding: 1.5rem;
    }

    .editable-field-group {
        margin-bottom: 1rem;
    }

    .editable-field-group label {
        font-weight: 500;
        color: #495057;
        display: block;
        margin-bottom: 0.3rem;
    }

    .editable-field {
        display: block; /* Делаем инпуты блочными */
        width: 100%; /* На всю ширину */
        padding: 0.5rem 0.75rem;
        font-size: 1rem;
        line-height: 1.5;
        color: #212529;
        background-color: #f8f9fa; /* Светлый фон для редактируемых полей */
        background-clip: padding-box;
        border: 1px solid #ced4da;
        border-radius: 0.25rem;
        transition: border-color .15s ease-in-out,box-shadow .15s ease-in-out;
    }
    .editable-field:focus {
        color: #212529;
        background-color: #fff;
        border-color: #86b7fe;
        outline: 0;
        box-shadow: 0 0 0 0.25rem rgba(13,110,253,.25);
    }
    .editable-field[readonly] {
        background-color: #e9ecef; /* Фон для нередактируемых */
        opacity: 1;
    }

    .missing-data-warning {
        background-color: #fff3cd;
        border-color: #ffecb5;
        color: #664d03;
        padding: 1rem;
        border-radius: 0.25rem;
    }
    .missing-data-warning ul {
        margin-bottom: 0;
        padding-left: 1.2rem;
    }

    .action-buttons {
        margin-top: 2rem;
        padding-top: 1rem;
        border-top: 1px solid #eee;
    }

    /* Стили для Карточки СИЗ */
    .ppe-card-section {
        background-color: #e9f5ff; /* Голубой фон */
        border-color: #b8daff;
    }
    .ppe-card-header {
        background-color: #cce5ff;
        color: #004085;
    }
    .missing-data-list {
        font-size: 0.9em;
        color: #dc3545; /* Красный для предупреждений */
    }
</style>
{% endblock %}

{% block content %}
<div class="container mt-4">
    <div class="row mb-3">
        <div class="col-12">
            <nav aria-label="breadcrumb">
                <ol class="breadcrumb">
                    <li class="breadcrumb-item"><a href="{% url 'directory:home' %}">Главная</a></li>
                    <li class="breadcrumb-item">
                        {% if employee %}
                            <a href="{% url 'directory:documents:document_selection' employee_id=employee.id %}">Выбор документов</a>
                        {% else %}
                            <span>Выбор документов</span>
                        {% endif %}
                    </li>
                    <li class="breadcrumb-item active" aria-current="page">{{ title }}</li>
                </ol>
            </nav>
        </div>
    </div>

    <div class="row justify-content-center">
        <div class="col-lg-10">
             {% if error_message %}
                <div class="alert alert-danger">{{ error_message }}</div>
             {% else %}
                <div class="card shadow-sm mb-4">
                    <div class="card-header bg-light">
                        <h4 class="mb-0"><i class="fas fa-user me-2"></i>Сотрудник: {{ employee.full_name_nominative }}</h4>
                    </div>
                    <div class="card-body">
                        <p class="mb-1"><strong>Должность:</strong> {{ employee.position.position_name|default:"-" }}</p>
                        <p class="mb-0"><strong>Организация:</strong> {{ employee.organization.short_name_ru|default:"-" }}</p>
                    </div>
                </div>

                {# Форма для отправки данных #}
                <form method="post" action="" id="documents-generation-form">
                    {% csrf_token %}

                    {# Блок для предпросмотра DOCX документов #}
                    {% if has_docx_documents %}
                        <h3 class="mb-3">Предпросмотр документов DOCX:</h3>
                        {% if docx_has_missing_data %}
                            <div class="alert alert-warning missing-data-warning mb-3">
                                <h5><i class="fas fa-exclamation-triangle me-2"></i>Внимание! Отсутствуют данные для DOCX:</h5>
                                <ul>
                                    {% for field_desc in docx_missing_data_fields %}
                                        <li>{{ field_desc }}</li>
                                    {% endfor %}
                                </ul>
                                <small>Некоторые поля помечены как "{{ MISSING_DATA_PLACEHOLDER }}". Пожалуйста, заполните их перед генерацией.</small>
                            </div>
                        {% endif %}

                        <div class="accordion" id="docxAccordion">
                        {% for doc_data in preview_data %}
                            <div class="accordion-item document-preview-section">
                                <h2 class="accordion-header" id="headingDocx{{ forloop.counter }}">
                                    <button class="accordion-button {% if doc_data.document_data.has_missing_data %}bg-warning{% endif %} {% if not forloop.first %}collapsed{% endif %}" type="button" data-bs-toggle="collapse" data-bs-target="#collapseDocx{{ forloop.counter }}" aria-expanded="{% if forloop.first %}true{% else %}false{% endif %}" aria-controls="collapseDocx{{ forloop.counter }}">
                                        <i class="fas fa-file-word me-2"></i>
                                        {{ document_types_dict|get_item:doc_data.document_type|default:doc_data.document_type }}
                                        {% if doc_data.document_data.has_missing_data %}<span class="badge bg-danger ms-2">Требует заполнения</span>{% endif %}
                                    </button>
                                </h2>
                                <div id="collapseDocx{{ forloop.counter }}" class="accordion-collapse collapse {% if forloop.first %}show{% endif %}" aria-labelledby="headingDocx{{ forloop.counter }}" data-bs-parent="#docxAccordion">
                                    <div class="accordion-body document-preview-body">
                                        {% if doc_data.document_data.has_missing_data %}
                                            <p class="text-danger mb-3"><strong>Пожалуйста, заполните поля, отмеченные как "{{ MISSING_DATA_PLACEHOLDER }}".</strong></p>
                                        {% endif %}
                                        <div class="row">
                                        {% for key, value in doc_data.document_data.items %}
                                            {% if key != 'missing_data_list' and key != 'has_missing_data' and key != 'employee_id' %}
                                                <div class="col-md-6 editable-field-group">
                                                    <label for="id_document_data_{{ doc_data.document_type }}_{{ key }}">{{ key|title }}:</label>
                                                    <input type="text"
                                                           id="id_document_data_{{ doc_data.document_type }}_{{ key }}"
                                                           class="form-control editable-field {% if value == MISSING_DATA_PLACEHOLDER %}is-invalid{% endif %}"
                                                           name="document_data_{{ doc_data.document_type }}_{{ key }}"
                                                           value="{{ value }}"
                                                           data-doc-type="{{ doc_data.document_type }}"
                                                           data-field-name="{{ key }}"
                                                           placeholder="Введите {{ key|lower }}">
                                                     {% if value == MISSING_DATA_PLACEHOLDER %}
                                                        <div class="invalid-feedback">Это поле необходимо заполнить.</div>
                                                    {% endif %}
                                                </div>
                                            {% endif %}
                                        {% endfor %}
                                        </div>
                                    </div>
                                </div>
                            </div>
                        {% endfor %}
                        </div>
                        {# Кнопка генерации DOCX #}
                        <div class="mt-3 mb-4 text-end">
                             <button type="submit" name="action" value="generate_docx" class="btn btn-success" {% if docx_has_missing_data %}disabled title="Заполните все обязательные поля"{% endif %}>
                                <i class="fas fa-file-word me-2"></i>Сгенерировать выбранные DOCX
                            </button>
                        </div>
                         <hr class="my-4">
                    {% endif %}

                    {# Блок для Карточки СИЗ #}
                    {% if generate_ppe_card %}
                        <div class="document-preview-section ppe-card-section">
                            <div class="document-preview-header ppe-card-header">
                                <h5><i class="fas fa-id-card me-2"></i>Карточка учета СИЗ (Excel)</h5>
                            </div>
                            <div class="document-preview-body">
                                <p>Будет сгенерирована личная карточка учета СИЗ в формате Excel для сотрудника.</p>
                                {% if ppe_card_has_missing_data %}
                                    <div class="alert alert-warning missing-data-warning">
                                        <strong>Внимание!</strong> Для полной Карточки СИЗ отсутствуют данные:
                                        <ul class="missing-data-list">
                                        {% for field_name in ppe_card_missing_data_fields %}
                                            <li>{{ field_name }}</li>
                                        {% endfor %}
                                        </ul>
                                        <small>Карточка будет сгенерирована без этих данных.</small>
                                    </div>
                                {% else %}
                                     <p class="text-success"><i class="fas fa-check-circle me-1"></i>Все необходимые данные для Карточки СИЗ найдены.</p>
                                {% endif %}
                                {# Кнопка генерации Excel #}
                                <div class="mt-3 text-end">
                                     <button type="submit" name="action" value="generate_ppe_card" class="btn btn-info">
                                        <i class="fas fa-file-excel me-2"></i>Сгенерировать Карточку СИЗ (Excel)
                                    </button>
                                </div>
                            </div>
                        </div>
                    {% endif %}

                    {# Общие кнопки действий #}
                    <div class="action-buttons d-flex justify-content-between">
                         <a href="{% url 'directory:documents:document_selection' employee_id=employee_id %}" class="btn btn-secondary">
                            <i class="fas fa-arrow-left me-2"></i>Назад к выбору
                        </a>
                        {# Кнопки генерации теперь отдельные для DOCX и Excel #}
                    </div>
                </form>
            {% endif %} {# end if not error_message #}
        </div>
    </div>
</div>
{% endblock %}

{% block extra_js %}
<script>
document.addEventListener('DOMContentLoaded', function() {
    // AJAX обновление данных документа при изменении полей DOCX
    const documentFields = document.querySelectorAll('.editable-field');
    documentFields.forEach(field => {
        field.addEventListener('input', function() { // Используем 'input' для немедленной реакции
            const docType = this.getAttribute('data-doc-type');
            const fieldName = this.getAttribute('data-field-name');
            const fieldValue = this.value.trim(); // Обрезаем пробелы

            // Снимаем класс is-invalid, если пользователь начал вводить
            if (fieldValue !== '') {
                this.classList.remove('is-invalid');
            }

            // Отправляем данные на сервер для обновления сессии (опционально, можно обновлять только перед отправкой)
            /* // Закомментировано, т.к. AJAX обновление может быть избыточным
            fetch('{% url "directory:documents:update_preview_data" %}', {
                method: 'POST',
                headers: {
                    'Content-Type': 'application/x-www-form-urlencoded',
                    'X-CSRFToken': document.querySelector('input[name="csrfmiddlewaretoken"]').value
                },
                body: `doc_type=${docType}&field_name=${fieldName}&field_value=${fieldValue}`
            })
            .then(response => response.json())
            .then(data => {
                if (!data.success) {
                    console.error('Ошибка при обновлении данных в сессии:', data.error);
                } else {
                     console.log(`Поле ${fieldName} для ${docType} обновлено в сессии.`);
                     // Можно добавить обновление статуса has_missing_data для кнопки, если нужно
                }
            })
            .catch(error => {
                console.error('Ошибка при отправке AJAX запроса:', error);
            });
            */
        });
    });

    // Валидация перед отправкой формы генерации DOCX
    const generateDocxButton = document.querySelector('button[name="action"][value="generate_docx"]');
    const form = document.getElementById('documents-generation-form');

    if (generateDocxButton && form) {
        form.addEventListener('submit', function(event) {
            // Проверяем только если нажата кнопка генерации DOCX
            if (document.activeElement === generateDocxButton) {
                let hasInvalidFields = false;
                const docxFields = form.querySelectorAll('.editable-field');
                docxFields.forEach(field => {
                    // Проверяем только поля видимых (развернутых) аккордеонов
                    const accordionItem = field.closest('.accordion-collapse');
                    if (accordionItem && accordionItem.classList.contains('show')) {
                         // Если поле пустое И оно было отмечено как недостающее (имеет is-invalid)
                        if (field.value.trim() === '' && field.classList.contains('is-invalid')) {
                            hasInvalidFields = true;
                             // Можно подсветить вкладку аккордеона
                             const headerButton = document.querySelector(`button[data-bs-target="#${accordionItem.id}"]`);
                             if(headerButton) headerButton.classList.add('border', 'border-danger');
                        } else {
                             // Убираем подсветку вкладки
                             const headerButton = document.querySelector(`button[data-bs-target="#${accordionItem.id}"]`);
                             if(headerButton) headerButton.classList.remove('border', 'border-danger');
                        }
                    }
                });

                if (hasInvalidFields) {
                    event.preventDefault(); // Останавливаем отправку формы
                    alert('Пожалуйста, заполните все обязательные поля DOCX, отмеченные красным.');
                    // Фокусируемся на первом невалидном поле
                    const firstInvalid = form.querySelector('.editable-field.is-invalid');
                    if(firstInvalid) firstInvalid.focus();
                }
            }
        });
    }


});
</script>
{% endblock %}
```

```python
# D:\YandexDisk\OT_online\directory\views\siz_issued.py
# 📁 directory/views/siz_issued.py
import re
import random
import logging
from django.views.generic import CreateView, DetailView, FormView, UpdateView, View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy, reverse
from django.shortcuts import get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect, FileResponse # Добавили FileResponse
from django.contrib import messages
from django.utils.decorators import method_decorator
from django.views.decorators.http import require_GET
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.template.loader import get_template
from io import BytesIO
from xhtml2pdf import pisa

from directory.models import Employee, SIZIssued, GeneratedDocument # Добавили GeneratedDocument
from directory.forms.siz_issued import SIZIssueForm, SIZIssueMassForm, SIZIssueReturnForm
from directory.utils.pdf import render_to_pdf
# Импортируем функцию генерации Excel
from directory.utils.excel_export import generate_card_excel

logger = logging.getLogger(__name__)

# --- Функции determine_gender_from_patronymic и get_random_siz_sizes без изменений ---
def determine_gender_from_patronymic(full_name):
    name_parts = full_name.split()
    if len(name_parts) >= 3: patronymic = name_parts[2]
    else: return "Мужской"
    if re.search(r'(ич|ыч)$', patronymic, re.IGNORECASE): return "Мужской"
    elif re.search(r'(на|вна|чна)$', patronymic, re.IGNORECASE): return "Женский"
    elif re.search(r'(оглы|улы|лы)$', patronymic, re.IGNORECASE): return "Мужской"
    elif re.search(r'(кызы|зы)$', patronymic, re.IGNORECASE): return "Женский"
    return "Мужской"

def get_random_siz_sizes(gender):
    if gender == "Мужской": headgear = random.randint(55, 59); gloves = random.randint(15, 19) / 2; respirator = random.choice(["1", "2", "3"])
    else: headgear = random.randint(53, 57); gloves = random.randint(13, 17) / 2; respirator = random.choice(["1", "2", "3"])
    gas_mask = respirator
    return {'headgear': headgear, 'gloves': gloves, 'respirator': respirator, 'gas_mask': gas_mask}

# --- SIZIssueFormView без изменений ---
class SIZIssueFormView(LoginRequiredMixin, CreateView):
    model = SIZIssued; form_class = SIZIssueForm; template_name = 'directory/siz_issued/issue_form.html'
    def get_success_url(self): return reverse('directory:siz:siz_personal_card', kwargs={'employee_id': self.object.employee.id})
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs(); kwargs['user'] = self.request.user
        employee_id = self.kwargs.get('employee_id');
        if employee_id: kwargs['employee_id'] = employee_id
        return kwargs
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs); context['title'] = 'Выдача СИЗ'
        employee_id = self.kwargs.get('employee_id')
        if employee_id:
            employee = get_object_or_404(Employee, id=employee_id); context['employee'] = employee
            if employee.position:
                from directory.models.siz import SIZNorm; norms = SIZNorm.objects.filter(position=employee.position).select_related('siz')
                context['base_norms'] = norms.filter(condition='')
                condition_groups = {};
                for norm in norms.exclude(condition=''):
                    if norm.condition not in condition_groups: condition_groups[norm.condition] = []
                    condition_groups[norm.condition].append(norm)
                context['condition_groups'] = [{'name': c, 'norms': n} for c, n in condition_groups.items()]
        return context
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f"✅ СИЗ '{self.object.siz.name}' успешно выдано сотруднику {self.object.employee.full_name_nominative}")
        return response

# --- issue_selected_siz без изменений ---
@login_required
def issue_selected_siz(request, employee_id):
    if request.method == 'POST':
        employee = get_object_or_404(Employee, id=employee_id)
        selected_norm_ids = request.POST.getlist('selected_norms')
        if not selected_norm_ids: messages.warning(request, "Не выбрано ни одного СИЗ для выдачи"); return redirect('directory:siz:siz_personal_card', employee_id=employee_id)
        from directory.models.siz import SIZNorm
        norms = SIZNorm.objects.filter(id__in=selected_norm_ids).select_related('siz')
        issued_count = 0
        for norm in norms:
            existing_issued = SIZIssued.objects.filter(employee=employee, siz=norm.siz, is_returned=False).exists()
            if not existing_issued:
                SIZIssued.objects.create(employee=employee, siz=norm.siz, quantity=norm.quantity, issue_date=timezone.now().date(), condition=norm.condition, received_signature=True); issued_count += 1
        if issued_count > 0: messages.success(request, f"✅ Успешно выдано {issued_count} наименований СИЗ сотруднику {employee.full_name_nominative}")
        else: messages.info(request, "ℹ️ Ни одно СИЗ не было выдано. Возможно, выбранные СИЗ уже находятся в использовании.")
    return redirect('directory:siz:siz_personal_card', employee_id=employee_id)

# --- SIZPersonalCardView без изменений ---
class SIZPersonalCardView(LoginRequiredMixin, DetailView):
    model = Employee; template_name = 'directory/siz_issued/personal_card.html'; context_object_name = 'employee'
    def get_object(self): return get_object_or_404(Employee, id=self.kwargs.get('employee_id'))
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs); context['title'] = f'Личная карточка учета СИЗ - {self.object.full_name_nominative}'
        issued_items = SIZIssued.objects.filter(employee=self.object).select_related('siz').order_by('-issue_date'); context['issued_items'] = issued_items
        if self.object.position:
            from directory.models.siz import SIZNorm; norms = SIZNorm.objects.filter(position=self.object.position).select_related('siz')
            context['base_norms'] = norms.filter(condition=''); conditions = list(set(norm.condition for norm in norms if norm.condition)); condition_groups = []
            for condition in conditions:
                condition_norms = [norm for norm in norms if norm.condition == condition]
                if condition_norms: condition_groups.append({'name': condition, 'norms': condition_norms})
            context['condition_groups'] = condition_groups
        gender = determine_gender_from_patronymic(self.object.full_name_nominative); context['gender'] = gender
        context['siz_sizes'] = get_random_siz_sizes(gender)
        return context

# --- SIZIssueReturnView без изменений ---
class SIZIssueReturnView(LoginRequiredMixin, UpdateView):
    model = SIZIssued; form_class = SIZIssueReturnForm; template_name = 'directory/siz_issued/return_form.html'; pk_url_kwarg = 'siz_issued_id'
    def get_success_url(self): return reverse('directory:siz:siz_personal_card', kwargs={'employee_id': self.object.employee.id})
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs); context['title'] = 'Возврат СИЗ'; context['employee'] = self.object.employee
        context['siz_name'] = self.object.siz.name; context['issue_date'] = self.object.issue_date
        return context
    def form_valid(self, form):
        response = super().form_valid(form); messages.success(self.request, f"✅ СИЗ '{self.object.siz.name}' успешно возвращено")
        return response

# --- employee_siz_issued_list без изменений ---
@login_required
@require_GET
def employee_siz_issued_list(request, employee_id):
    employee = get_object_or_404(Employee, pk=employee_id)
    issued_items = SIZIssued.objects.filter(employee=employee).select_related('siz').order_by('-issue_date')
    result = {'employee_id': employee.id, 'employee_name': employee.full_name_nominative, 'position': employee.position.position_name if employee.position else "", 'organization': employee.organization.short_name_ru, 'issued_items': []}
    for item in issued_items:
        item_data = {'id': item.id, 'siz_name': item.siz.name, 'siz_classification': item.siz.classification, 'issue_date': item.issue_date.strftime('%d.%m.%Y'), 'quantity': item.quantity, 'wear_percentage': item.wear_percentage, 'is_returned': item.is_returned, 'return_date': item.return_date.strftime('%d.%m.%Y') if item.return_date else None, 'notes': item.notes, 'condition': item.condition}
        result['issued_items'].append(item_data)
    return JsonResponse(result)

# --- export_personal_card_pdf без изменений ---
def export_personal_card_pdf(request, employee_id):
    employee = get_object_or_404(Employee, pk=employee_id)
    issued_items = SIZIssued.objects.filter(employee=employee).select_related('siz').order_by('-issue_date')
    selected_norm_ids = request.GET.getlist('selected_norms')
    if not selected_norm_ids and employee.position:
        from directory.models.siz import SIZNorm
        all_norms = SIZNorm.objects.filter(position=employee.position).values_list('id', flat=True)
        selected_norm_ids = list(map(str, all_norms))
    selected_items = []
    if selected_norm_ids:
        from directory.models.siz import SIZNorm
        selected_norms = SIZNorm.objects.filter(id__in=selected_norm_ids).select_related('siz')
        for norm in selected_norms: selected_items.append({'siz': norm.siz, 'classification': norm.siz.classification, 'quantity': norm.quantity})
    base_norms = []; condition_groups = []
    if employee.position:
        from directory.models.siz import SIZNorm
        norms = SIZNorm.objects.filter(position=employee.position).select_related('siz')
        base_norms = list(norms.filter(condition=''))
        conditions = list(set(norm.condition for norm in norms if norm.condition))
        for condition in conditions:
            condition_norms = [norm for norm in norms if norm.condition == condition]
            if condition_norms: condition_groups.append({'name': condition, 'norms': condition_norms})
    context = {'employee': employee, 'issued_items': issued_items, 'base_norms': base_norms, 'condition_groups': condition_groups, 'today': timezone.now().date(), 'gender': determine_gender_from_patronymic(employee.full_name_nominative), 'siz_sizes': get_random_siz_sizes(determine_gender_from_patronymic(employee.full_name_nominative)), 'selected_items': selected_items}
    filename = f"personal_card_{employee.full_name_nominative.replace(' ', '_')}.pdf"
    template_path = 'directory/siz_issued/personal_card_pdf_landscape.html'
    try:
        pdf_options = {'page-size': 'A4', 'margin-top': '0.5cm', 'margin-right': '0.5cm', 'margin-bottom': '0.5cm', 'margin-left': '0.5cm', 'encoding': "UTF-8"}
        return render_to_pdf(template_path=template_path, context=context, filename=filename, as_attachment=True, pdf_options=pdf_options, landscape=True) # Добавили landscape=True
    except Exception as e:
        logger.error(f"Ошибка при создании PDF: {e}"); messages.error(request, f"Ошибка при создании PDF: {e}")
        return redirect('directory:siz:siz_personal_card', employee_id=employee_id)

# --- Представление для генерации Excel из URL ---
@login_required
def export_personal_card_excel_view(request, employee_id):
    """
    Представление для генерации и скачивания Excel-карточки СИЗ по URL.
    Использует существующую функцию generate_card_excel.
    """
    logger.info(f"Запрос на генерацию Excel-карточки для сотрудника {employee_id} через URL")
    # Вызываем функцию генерации Excel, НЕ сохраняя в БД
    response = generate_card_excel(request, employee_id, save_to_db=False, user=request.user)

    # generate_card_excel вернет либо FileResponse, либо редирект с сообщением об ошибке
    return response
""",
}

# --- ЛОГИКА СКРИПТА ---
def update_project_files():
    """
    Перезаписывает указанные файлы в проекте новым содержимым.
    """
    print("--- НАЧАЛО ОБНОВЛЕНИЯ ФАЙЛОВ ПРОЕКТА ---")
    print(f"Корневая папка проекта: {PROJECT_ROOT}")

    if not os.path.isdir(PROJECT_ROOT):
        print(f"\nОШИБКА: Корневая папка проекта не найдена: {PROJECT_ROOT}")
        print("Пожалуйста, проверьте и исправьте путь в переменной PROJECT_ROOT.")
        sys.exit(1)

    updated_count = 0
    error_count = 0

    for relative_path, new_content in UPDATED_FILES.items():
        # Нормализуем относительный путь для текущей ОС
        normalized_relative_path = os.path.normpath(relative_path)
        absolute_path = os.path.join(PROJECT_ROOT, normalized_relative_path)

        print(f"\nОбработка файла: {normalized_relative_path}")
        print(f"Полный путь: {absolute_path}")

        try:
            # --- Создание директорий, если их нет ---
            directory_path = os.path.dirname(absolute_path)
            if not os.path.exists(directory_path):
                print(f"Создание директории: {directory_path}")
                os.makedirs(directory_path, exist_ok=True)
            elif not os.path.isdir(directory_path):
                 print(f"ОШИБКА: Путь для директории '{directory_path}' занят файлом!")
                 error_count += 1
                 continue # Переходим к следующему файлу

            # --- Перезапись файла ---
            # Убираем возможный лишний перенос строки в начале контента
            content_to_write = new_content.strip()

            with open(absolute_path, 'w', encoding='utf-8') as f:
                f.write(content_to_write)
            print(f"УСПЕШНО: Файл '{normalized_relative_path}' обновлен.")
            updated_count += 1

        except IOError as e:
            print(f"ОШИБКА ВВОДА/ВЫВОДА при обновлении файла '{normalized_relative_path}': {e}")
            error_count += 1
        except PermissionError as e:
            print(f"ОШИБКА ДОСТУПА при обновлении файла '{normalized_relative_path}': {e}")
            print("Проверьте права на запись.")
            error_count += 1
        except Exception as e:
            print(f"НЕПРЕДВИДЕННАЯ ОШИБКА при обновлении файла '{normalized_relative_path}': {e}")
            error_count += 1

    print("\n--- ЗАВЕРШЕНИЕ ОБНОВЛЕНИЯ ФАЙЛОВ ---")
    print(f"Всего обработано файлов: {len(UPDATED_FILES)}")
    print(f"Успешно обновлено: {updated_count}")
    print(f"Ошибок: {error_count}")

    if error_count > 0:
        print("\nВНИМАНИЕ: Во время обновления произошли ошибки. Просмотрите лог выше.")
    else:
        print("\nВсе указанные файлы успешно обновлены.")

if __name__ == "__main__":
    # Запрос подтверждения у пользователя
    confirm = input(f"Вы уверены, что хотите перезаписать файлы в директории '{PROJECT_ROOT}'? (yes/no): ").lower()
    if confirm == 'yes':
        update_project_files()
    else:
        print("Обновление отменено.")